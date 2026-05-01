"""Tests for the Merchant Charges module.

Covers loader validation + drop accounting, brand normalization, partition
+ sort logic, the data sheet (Credits & Refunds + dual grand totals), and
the Summary dashboard sheet.
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from app.modules.merchant_charges.services import (
    CREDITS_SECTION_LABEL,
    DROP_REASON_BAD_AMOUNT,
    DROP_REASON_BLANK_MERCHANT,
    OUTPUT_SHEET_NAME,
    SUMMARY_SHEET_NAME,
    MerchantChargesError,
    _brand_of,
    _pick_display_brand,
    _to_match_key,
    group_and_sort,
    load_transactions,
    load_transactions_with_stats,
    process_merchant_charges,
    write_report,
)

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "merchant_charges_sample.xlsx"
CUSTOMER_INPUT = Path(
    "/Users/jordanhill/Desktop/Baseline-future-sit-down-and-rearage-from-cowork/"
    "Baseline-Ops-Supporting-docs-older/RCwendt-Colissa/Merchant-Charges/Input.xlsx"
)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _build_workbook(headers: list[str], rows: list[list]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _full_headers() -> list[str]:
    return [
        "Posted", "Occurred", "Merchant Name", "Merchant City", "Merchant State",
        "Merchant Zip Code", "MCC/SIC Code", "MCC Description", "Original Amount",
        "Currency Desc", "Conversion Rate", "Billed Amount", "Memo", "Debit Credit",
        "Reference Number", "Statement Cycle", "Account Name", "Account Number",
    ]


def _row(posted, occurred, merchant: str, billed) -> list:
    return [posted, occurred, merchant, "", "", "", "", "", billed, "", "", billed,
            "", "", "", None, "", ""]


def _build_input(rows: list[tuple]) -> bytes:
    """Each `rows` entry is (posted, occurred, merchant, billed)."""
    return _build_workbook(_full_headers(), [_row(*r) for r in rows])


def _eval_sum_range(ws, formula: str) -> float:
    """Evaluate a literal '=SUM(D{a}:D{b})' against the cached cell values."""
    assert formula.startswith("=SUM(") and formula.endswith(")")
    inner = formula[len("=SUM("):-1]
    start, end = inner.split(":")
    col_letter = start[0]
    col_idx = openpyxl.utils.column_index_from_string(col_letter)
    r0 = int(start[1:])
    r1 = int(end[1:])
    total = 0.0
    for r in range(r0, r1 + 1):
        total += float(ws.cell(row=r, column=col_idx).value)
    return total


def _eval_brand_subtotals(ws, charges_formula: str) -> float:
    """Evaluate '=SUM(E{a},E{b},...)' by walking each referenced cell to the
    underlying =SUM(D{}:D{}) brand subtotal and summing the raw amounts.
    """
    assert charges_formula.startswith("=SUM(") and charges_formula.endswith(")")
    refs = charges_formula[len("=SUM("):-1].split(",")
    total = 0.0
    for ref in refs:
        sub_row = int(ref.strip()[1:])
        sub_formula = ws.cell(row=sub_row, column=4).value
        total += _eval_sum_range(ws, sub_formula)
    return total


def _data_ws(wb_bytes: bytes):
    """Open the workbook and return the Merchant Charges sheet (formula view)."""
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)[OUTPUT_SHEET_NAME]


def _summary_ws(wb_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)[SUMMARY_SHEET_NAME]


def _find_grand_total_rows(ws) -> dict[str, int]:
    out: dict[str, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=3).value
        if isinstance(v, str) and v.startswith("Grand Total"):
            out[v] = r
    return out


def _find_credits_header_row(ws) -> int | None:
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=3).value == CREDITS_SECTION_LABEL:
            return r
    return None


def _summary_value_for(ws, label: str):
    """Find row where col B == label, return col C value."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == label:
            return ws.cell(row=r, column=3).value
    return None


def _summary_has_label(ws, label: str) -> bool:
    return any(ws.cell(row=r, column=2).value == label for r in range(1, ws.max_row + 1))


# ---------------------------------------------------------------------------
# loader
# ---------------------------------------------------------------------------

class TestLoadTransactions:
    def test_happy_path(self):
        data = _build_input([(datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97)])
        txs = load_transactions(data)
        assert len(txs) == 1
        assert set(txs[0].keys()) == {"posted", "occurred", "merchant", "amount"}
        assert txs[0]["amount"] == 47.97

    def test_missing_required_column(self):
        bad = ["Posted", "Occurred", "Merchant Name"]
        data = _build_workbook(bad, [[datetime(2026, 3, 5), datetime(2026, 3, 4), "X"]])
        with pytest.raises(MerchantChargesError, match="Billed Amount"):
            load_transactions(data)

    def test_drops_blank_merchant(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 10.0),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "", 99.0),
        ])
        txs = load_transactions(data)
        assert [t["merchant"] for t in txs] == ["ARCO"]

    def test_drops_unparseable_amount(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 10.0),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "BADAMT", "not-a-number"),
        ])
        txs = load_transactions(data)
        assert [t["merchant"] for t in txs] == ["ARCO"]

    def test_no_valid_rows_raises(self):
        data = _build_workbook(_full_headers(), [])
        with pytest.raises(MerchantChargesError, match="No valid transactions"):
            load_transactions(data)

    def test_header_lookup_is_case_insensitive(self):
        headers = _full_headers()
        headers[2] = "  merchant  name  "
        headers[11] = "BILLED AMOUNT"
        data = _build_workbook(
            headers,
            [_row(datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 10.0)],
        )
        txs = load_transactions(data)
        assert txs[0]["merchant"] == "ARCO"


class TestLoadStats:
    def test_no_drops_yields_empty_drop_counts(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 10.0),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "ALBERTSONS", 20.0),
        ])
        txs, stats = load_transactions_with_stats(data)
        assert stats.total_input_rows == 2
        assert stats.drop_counts == {}
        assert len(txs) == 2

    def test_drop_counts_track_each_reason(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 10.0),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "", 99.0),                # blank
            (datetime(2026, 3, 7), datetime(2026, 3, 6), "BADAMT", "not-a-number"),  # bad amount
            (datetime(2026, 3, 8), datetime(2026, 3, 7), "BADAMT2", "also-bad"),     # bad amount
        ])
        txs, stats = load_transactions_with_stats(data)
        assert stats.total_input_rows == 4
        assert stats.drop_counts == {DROP_REASON_BLANK_MERCHANT: 1, DROP_REASON_BAD_AMOUNT: 2}
        assert len(txs) == 1


# ---------------------------------------------------------------------------
# brand normalization
# ---------------------------------------------------------------------------

class TestBrandOf:
    @pytest.mark.parametrize("merchant,expected", [
        ("ARCO 912987", "ARCO"),
        ("ARCO 913712", "ARCO"),
        ("7-ELEVEN 20342", "7-ELEVEN"),
        ("THE HOME DEPOT #0668", "THE HOME DEPOT"),
        ("THE HOME DEPOT 673", "THE HOME DEPOT"),
        ("STATERBROS114", "STATERBROS"),
        ("SHERWIN-WILLIAMS701542", "SHERWIN-WILLIAMS"),
        ("LOWES #01574*", "LOWES"),
        ("DUNN-EDWARDS # 170", "DUNN-EDWARDS"),
        ("DUNN-EDWARDS CORP #146", "DUNN-EDWARDS CORP"),
        ("ALBERTSONS #0708", "ALBERTSONS"),
        ("WM SUPERCENTER #3516", "WM SUPERCENTER"),
        ("CHEVRON 0383260", "CHEVRON"),
        ("THE UPS STORE 6987", "THE UPS STORE"),
        ("CORONA OIL", "CORONA OIL"),
        ("D B SERVICE CENTER", "D B SERVICE CENTER"),
        ("76 - TEMECULA FUELS CORP", "76 - TEMECULA FUELS CORP"),
        ("STARS & STRIPES CHULA VI", "STARS & STRIPES CHULA VI"),
        ("VISTA PAINT 009 - SAN D", "VISTA PAINT 009 - SAN D"),
        ("PAYMENT RECEIVED - THANK", "PAYMENT RECEIVED - THANK"),
    ])
    def test_brand_extraction(self, merchant: str, expected: str):
        assert _brand_of(merchant) == expected


# ---------------------------------------------------------------------------
# group_and_sort
# ---------------------------------------------------------------------------

class TestGroupAndSort:
    def test_returns_tuple_with_positives_and_negatives(self):
        txs = load_transactions(FIXTURE_PATH.read_bytes())
        positives, negatives = group_and_sort(txs)
        assert sum(len(r) for r in positives.values()) + len(negatives) == len(txs)

    def test_brand_keys_alphabetical_excel_style(self):
        txs = [
            {"posted": datetime(2026, 3, 5), "occurred": datetime(2026, 3, 4),
             "merchant": "7-ELEVEN 33252", "amount": 10.0},
            {"posted": datetime(2026, 3, 5), "occurred": datetime(2026, 3, 4),
             "merchant": "76 - TEMECULA FUELS CORP", "amount": 20.0},
            {"posted": datetime(2026, 3, 5), "occurred": datetime(2026, 3, 4),
             "merchant": "ALBERTSONS #0708", "amount": 30.0},
        ]
        positives, negatives = group_and_sort(txs)
        assert list(positives.keys()) == ["76 - TEMECULA FUELS CORP", "7-ELEVEN", "ALBERTSONS"]
        assert negatives == []

    def test_within_brand_sort_by_merchant_then_posted_then_occurred(self):
        txs = [
            {"posted": datetime(2026, 3, 5), "occurred": datetime(2026, 3, 4),
             "merchant": "THE HOME DEPOT #1018", "amount": 40.73},
            {"posted": datetime(2026, 3, 12), "occurred": datetime(2026, 3, 10),
             "merchant": "THE HOME DEPOT #0668", "amount": 15.93},
            {"posted": datetime(2026, 3, 13), "occurred": datetime(2026, 3, 11),
             "merchant": "THE HOME DEPOT #0668", "amount": 99.99},
        ]
        positives, _ = group_and_sort(txs)
        rows = positives["THE HOME DEPOT"]
        assert [r["merchant"] for r in rows] == [
            "THE HOME DEPOT #0668", "THE HOME DEPOT #0668", "THE HOME DEPOT #1018",
        ]
        assert rows[0]["posted"] < rows[1]["posted"]


# ---------------------------------------------------------------------------
# data-sheet basics (Merchant Charges sheet — wb.active is now Summary, so
# tests have to look up the data sheet by name)
# ---------------------------------------------------------------------------

class TestDataSheetBasics:
    def _fixture(self):
        return group_and_sort(load_transactions(FIXTURE_PATH.read_bytes()))

    def test_data_sheet_title_and_headers(self):
        positives, negatives = self._fixture()
        wb = openpyxl.load_workbook(io.BytesIO(write_report(positives, negatives)))
        ws = wb[OUTPUT_SHEET_NAME]
        assert ws.title == OUTPUT_SHEET_NAME
        assert [c.value for c in ws[1]] == list(
            ("Posted Date", "Occurred Date", "Merchant Name", "Amount", "Merchant Total")
        )

    def test_freeze_pane_set_on_data_sheet(self):
        positives, negatives = self._fixture()
        wb = openpyxl.load_workbook(io.BytesIO(write_report(positives, negatives)))
        assert wb[OUTPUT_SHEET_NAME].freeze_panes == "A2"

    def test_brand_subtotals_use_formulas(self):
        positives, negatives = self._fixture()
        ws = _data_ws(write_report(positives, negatives))
        formula_count = sum(
            1 for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=4).value, str)
            and ws.cell(row=r, column=4).value.startswith("=SUM(D")
        )
        assert formula_count >= len(positives)


# ---------------------------------------------------------------------------
# negatives section + dual grand totals
# ---------------------------------------------------------------------------

class TestNegativesSection:
    def test_negative_transactions_moved_to_credits_section(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "ARCO 913712", 50.90),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "THE HOME DEPOT #0668", 15.93),
            (datetime(2026, 3, 27), datetime(2026, 3, 27), "PAYMENT RECEIVED", -100.00),
        ])
        ws = _data_ws(process_merchant_charges(data))

        sum_d_count = sum(
            1 for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=4).value, str)
            and ws.cell(row=r, column=4).value.startswith("=SUM(D")
        )
        # 2 brand subtotals (ARCO + THE HOME DEPOT) + 1 credits subtotal.
        assert sum_d_count == 3

        credits_row = _find_credits_header_row(ws)
        assert credits_row is not None
        assert ws.cell(row=credits_row + 1, column=3).value == "PAYMENT RECEIVED"
        sub_row = credits_row + 2
        assert ws.cell(row=sub_row, column=4).value.startswith("=SUM(D")
        assert ws.cell(row=sub_row, column=5).value == ws.cell(row=sub_row, column=4).value

        gts = _find_grand_total_rows(ws)
        assert "Grand Total (Charges Only)" in gts
        assert "Grand Total (Net)" in gts

    def test_no_negatives_renders_single_grand_total(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "THE HOME DEPOT #0668", 15.93),
        ])
        ws = _data_ws(process_merchant_charges(data))
        assert _find_credits_header_row(ws) is None
        gts = _find_grand_total_rows(ws)
        assert list(gts.keys()) == ["Grand Total"]
        formula = ws.cell(row=gts["Grand Total"], column=5).value
        assert formula.startswith("=SUM(E")

    def test_brand_with_mixed_signs_splits_correctly(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 100.00),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "ARCO 913712", -30.00),
        ])
        positives, negatives = group_and_sort(load_transactions(data))
        assert "ARCO" in positives
        assert len(positives["ARCO"]) == 1
        assert positives["ARCO"][0]["amount"] == 100.0
        assert len(negatives) == 1 and negatives[0]["amount"] == -30.0

        ws = _data_ws(write_report(positives, negatives))
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=4).value
            if isinstance(v, str) and v.startswith("=SUM(D"):
                assert _eval_sum_range(ws, v) in {100.0, -30.0}

        gts = _find_grand_total_rows(ws)
        charges_formula = ws.cell(row=gts["Grand Total (Charges Only)"], column=5).value
        assert abs(_eval_brand_subtotals(ws, charges_formula) - 100.0) < 0.01

        net_formula = ws.cell(row=gts["Grand Total (Net)"], column=5).value
        assert net_formula.startswith("=E") and "+" in net_formula
        assert "SUM" not in net_formula

    def test_credits_section_sort_order(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 10.0),
            (datetime(2026, 3, 20), datetime(2026, 3, 18), "ZZZ LATE",  -1.00),
            (datetime(2026, 3, 5), datetime(2026, 3, 3), "BBB",         -100.00),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "AAA",         -50.00),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "AAB",         -2.00),
        ])
        ws = _data_ws(process_merchant_charges(data))
        hdr = _find_credits_header_row(ws)
        expected = ["BBB", "AAA", "AAB", "ZZZ LATE"]
        actual = [ws.cell(row=hdr + 1 + i, column=3).value for i in range(len(expected))]
        assert actual == expected

    def test_credits_section_subtotal_formula_is_negative_sum(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO", 100.0),
            (datetime(2026, 3, 27), datetime(2026, 3, 27), "PAYMENT RECEIVED", -500.00),
            (datetime(2026, 4, 1), datetime(2026, 3, 31), "REFUND HOME DEPOT", -25.50),
        ])
        ws = _data_ws(process_merchant_charges(data))
        hdr = _find_credits_header_row(ws)
        sub_row = None
        for r in range(hdr + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=4).value
            if isinstance(v, str) and v.startswith("=SUM(D"):
                sub_row = r
                break
        assert sub_row is not None
        sub_formula = ws.cell(row=sub_row, column=4).value
        computed = _eval_sum_range(ws, sub_formula)
        assert computed < 0
        assert abs(computed - (-525.50)) < 0.01

    def test_grand_total_net_equals_charges_plus_credits(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "ARCO 913712", 103.29),
            (datetime(2026, 3, 18), datetime(2026, 3, 16), "THE HOME DEPOT #0668", 15.93),
            (datetime(2026, 3, 27), datetime(2026, 3, 27), "PAYMENT RECEIVED", -200.00),
        ])
        ws = _data_ws(process_merchant_charges(data))
        gts = _find_grand_total_rows(ws)
        charges = _eval_brand_subtotals(
            ws, ws.cell(row=gts["Grand Total (Charges Only)"], column=5).value
        )
        net_formula = ws.cell(row=gts["Grand Total (Net)"], column=5).value
        parts = net_formula.lstrip("=").split("+")
        credits_row = int(parts[1].strip()[1:])
        credits = _eval_sum_range(ws, ws.cell(row=credits_row, column=4).value)
        net = charges + credits
        all_amounts = sum(t["amount"] for t in load_transactions(data))
        assert abs(net - all_amounts) < 0.01

    def test_grand_total_charges_only_excludes_negatives(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "ARCO 913712", 103.29),
            (datetime(2026, 3, 27), datetime(2026, 3, 27), "PAYMENT RECEIVED", -500.00),
        ])
        ws = _data_ws(process_merchant_charges(data))
        gts = _find_grand_total_rows(ws)
        charges = _eval_brand_subtotals(
            ws, ws.cell(row=gts["Grand Total (Charges Only)"], column=5).value
        )
        assert abs(charges - (47.97 + 103.29)) < 0.01

    def test_all_negatives_renders_only_net_grand_total(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "REFUND A", -10.0),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "REFUND B", -20.0),
        ])
        ws = _data_ws(process_merchant_charges(data))
        gts = _find_grand_total_rows(ws)
        assert "Grand Total (Charges Only)" not in gts
        assert "Grand Total (Net)" in gts


# ---------------------------------------------------------------------------
# Summary sheet (the new dashboard)
# ---------------------------------------------------------------------------

class TestSummarySheet:
    def test_output_has_two_sheets_in_correct_order(self):
        out = process_merchant_charges(FIXTURE_PATH.read_bytes())
        wb = openpyxl.load_workbook(io.BytesIO(out))
        assert wb.sheetnames == [SUMMARY_SHEET_NAME, OUTPUT_SHEET_NAME]

    def test_summary_is_active_sheet(self):
        out = process_merchant_charges(FIXTURE_PATH.read_bytes())
        wb = openpyxl.load_workbook(io.BytesIO(out))
        assert wb.active.title == SUMMARY_SHEET_NAME

    def test_summary_sheet_counts_match_input(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 10.0),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ALBERTSONS", 20.0),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 30.0),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "CHEVRON 9999", 40.0),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "LOWES #01574*", 50.0),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "REFUND A", -1.0),
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "REFUND B", -2.0),
        ])
        ws = _summary_ws(process_merchant_charges(data))
        assert _summary_value_for(ws, "Total transactions in input:") == 7
        assert _summary_value_for(ws, "Charges (positive):") == 5
        assert _summary_value_for(ws, "Credits (negative):") == 2
        assert not _summary_has_label(ws, "Rows dropped (validation):")

    def test_summary_dollar_formulas_reference_correct_cells(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "ARCO 913712", 103.29),
            (datetime(2026, 3, 27), datetime(2026, 3, 27), "PAYMENT RECEIVED", -200.00),
        ])
        out = process_merchant_charges(data)
        data_ws = _data_ws(out)
        sum_ws = _summary_ws(out)

        # Find the Grand Total (Charges Only) row on the data sheet, then
        # confirm the Summary sheet's formula points at that exact cell.
        gts = _find_grand_total_rows(data_ws)
        charges_row = gts["Grand Total (Charges Only)"]
        net_row = gts["Grand Total (Net)"]

        charges_formula = _summary_value_for(sum_ws, "Grand Total (Charges Only):")
        assert charges_formula == f"='{OUTPUT_SHEET_NAME}'!E{charges_row}"

        net_formula = _summary_value_for(sum_ws, "Grand Total (Net):")
        assert net_formula == f"='{OUTPUT_SHEET_NAME}'!E{net_row}"

        # Credits subtotal: locate the row where the data sheet has =SUM(D...)
        # for the credits section (row right after "Credits & Refunds" + neg rows).
        hdr = _find_credits_header_row(data_ws)
        credits_sub_row = None
        for r in range(hdr + 1, data_ws.max_row + 1):
            v = data_ws.cell(row=r, column=4).value
            if isinstance(v, str) and v.startswith("=SUM(D"):
                credits_sub_row = r
                break
        credits_formula = _summary_value_for(sum_ws, "Credits & Refunds subtotal:")
        assert credits_formula == f"='{OUTPUT_SHEET_NAME}'!E{credits_sub_row}"

    def test_summary_formulas_recalculate_to_correct_values(self):
        # No project recalc helper exists; resolve cross-sheet formulas in
        # Python by walking back to the underlying data-sheet cells.
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "ARCO 913712", 103.29),
            (datetime(2026, 3, 27), datetime(2026, 3, 27), "PAYMENT RECEIVED", -200.00),
        ])
        out = process_merchant_charges(data)
        data_ws = _data_ws(out)
        sum_ws = _summary_ws(out)

        charges_formula = _summary_value_for(sum_ws, "Grand Total (Charges Only):")
        # Format: ='Merchant Charges'!E{row}
        ref_row = int(charges_formula.split("!E")[1])
        data_charges = _eval_brand_subtotals(data_ws, data_ws.cell(row=ref_row, column=5).value)
        expected = sum(t["amount"] for t in load_transactions(data) if t["amount"] >= 0)
        assert abs(data_charges - expected) < 0.01

    def test_summary_omits_credit_rows_when_no_negatives(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 12), datetime(2026, 3, 10), "ALBERTSONS", 25.0),
        ])
        ws = _summary_ws(process_merchant_charges(data))
        assert not _summary_has_label(ws, "Credits & Refunds subtotal:")
        assert not _summary_has_label(ws, "Grand Total (Net):")
        assert _summary_has_label(ws, "Grand Total (Charges Only):")

    def test_summary_omits_dropped_section_when_no_drops(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
        ])
        ws = _summary_ws(process_merchant_charges(data))
        assert not _summary_has_label(ws, "Rows dropped (validation):")

    def test_summary_includes_dropped_section_with_per_reason_breakdown(self):
        data = _build_input([
            (datetime(2026, 3, 5), datetime(2026, 3, 4), "ARCO 912987", 47.97),
            (datetime(2026, 3, 6), datetime(2026, 3, 5), "", 99.0),                # blank merchant
            (datetime(2026, 3, 7), datetime(2026, 3, 6), "BAD", "not-a-number"),    # bad amount
            (datetime(2026, 3, 8), datetime(2026, 3, 7), "BAD2", "still-bad"),      # bad amount
        ])
        ws = _summary_ws(process_merchant_charges(data))
        assert _summary_value_for(ws, "Rows dropped (validation):") == 3
        assert _summary_value_for(ws, f"   {DROP_REASON_BLANK_MERCHANT}:") == 1
        assert _summary_value_for(ws, f"   {DROP_REASON_BAD_AMOUNT}:") == 2

    def test_summary_date_range_uses_min_max_posted(self):
        data = _build_input([
            (datetime(2026, 4, 30), datetime(2026, 4, 28), "ARCO", 10.0),
            (datetime(2026, 3, 4), datetime(2026, 3, 3), "ALBERTSONS", 20.0),
            (datetime(2026, 3, 15), datetime(2026, 3, 14), "CHEVRON 9", 5.0),
        ])
        ws = _summary_ws(process_merchant_charges(data))
        date_range = _summary_value_for(ws, "Date range (Posted):")
        assert date_range == "3/4/2026 \u2013 4/30/2026"

    def test_merchant_charges_sheet_unchanged_by_summary_addition(self):
        """Regression guard: the data sheet's content must match what an
        equivalent legacy (pre-Summary) writer would have produced."""
        positives, negatives = group_and_sort(load_transactions(FIXTURE_PATH.read_bytes()))
        out = write_report(positives, negatives)
        ws = _data_ws(out)

        # Header row unchanged.
        assert [c.value for c in ws[1]] == list(
            ("Posted Date", "Occurred Date", "Merchant Name", "Amount", "Merchant Total")
        )
        # Row 2 is the first data row of the first brand — i.e., NO inline
        # summary block was inserted at the top of the data sheet.
        assert isinstance(ws.cell(row=2, column=1).value, datetime)
        # Freeze pane preserved.
        assert ws.freeze_panes == "A2"

    def test_summary_tab_is_visually_distinct(self):
        out = process_merchant_charges(FIXTURE_PATH.read_bytes())
        wb = openpyxl.load_workbook(io.BytesIO(out))
        sum_ws = wb[SUMMARY_SHEET_NAME]
        # Tab color set; gridlines hidden.
        assert sum_ws.sheet_properties.tabColor is not None
        assert sum_ws.sheet_view.showGridLines is False


# ---------------------------------------------------------------------------
# whitespace-insensitive brand merging
# ---------------------------------------------------------------------------

class TestMatchKeyHelper:
    @pytest.mark.parametrize("normalized,expected_key", [
        ("BURGER KING", "BURGERKING"),
        ("BURGERKING", "BURGERKING"),
        ("BURGER  KING", "BURGERKING"),  # multi-space (paranoid: should never happen post-normalize)
        ("THE HOME DEPOT", "THEHOMEDEPOT"),
        ("HOMEDEPOT", "HOMEDEPOT"),       # different match key from THE HOME DEPOT
        ("WAL-MART", "WAL-MART"),
        ("WALMART", "WALMART"),
        ("7-ELEVEN", "7-ELEVEN"),
        ("7 ELEVEN", "7ELEVEN"),
        ("BK", "BK"),
        ("STARBUCKS", "STARBUCKS"),
    ])
    def test_match_key_strips_whitespace_only(self, normalized: str, expected_key: str):
        assert _to_match_key(normalized) == expected_key


class TestPickDisplayBrand:
    def test_prefers_form_with_whitespace(self):
        # Single spaced occurrence beats five unspaced ones (rule 1 absolute).
        forms = ["BURGERKING", "BURGERKING", "BURGERKING", "BURGERKING", "BURGERKING", "BURGER KING"]
        assert _pick_display_brand(forms) == "BURGER KING"

    def test_most_frequent_among_spaced(self):
        forms = ["BURGER KING", "BURGER KING", "BURGER KING DELUXE", "BURGERKING"]
        assert _pick_display_brand(forms) == "BURGER KING"

    def test_alphabetical_tiebreaker_among_spaced(self):
        # Two spaced forms tied at one occurrence each — alphabetical wins.
        forms = ["B FOO", "A FOO"]
        assert _pick_display_brand(forms) == "A FOO"

    def test_falls_back_to_unspaced_when_no_spaced_form(self):
        forms = ["BURGERKING", "BURGERKING", "BURGERKING"]
        assert _pick_display_brand(forms) == "BURGERKING"


class TestBrandOfMultiSpace:
    """`_brand_of` must collapse internal whitespace runs."""

    def test_double_space_collapsed(self):
        assert _brand_of("BURGER  KING 09812") == "BURGER KING"

    def test_tab_or_mixed_whitespace_collapsed(self):
        assert _brand_of("BURGER\tKING #4421") == "BURGER KING"

    def test_leading_and_trailing_whitespace_stripped(self):
        assert _brand_of("   BURGER KING #4421   ") == "BURGER KING"


class TestWhitespaceInsensitiveMerging:
    def test_whitespace_only_difference_merges(self):
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGER KING #4421", 9.85),
            (datetime(2026, 3, 8), datetime(2026, 3, 8), "BURGERKING #2231", 8.95),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        assert list(positives.keys()) == ["BURGER KING"]
        rows = positives["BURGER KING"]
        assert len(rows) == 2
        assert abs(sum(r["amount"] for r in rows) - 18.80) < 0.01

    def test_multiple_whitespace_variants_merge(self):
        # Three forms: single-space, no-space, double-space. The double-space
        # variant should already be cleaned to single-space by _brand_of.
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGER KING #4421", 9.85),
            (datetime(2026, 3, 8), datetime(2026, 3, 8), "BURGERKING #2231", 8.95),
            (datetime(2026, 3, 5), datetime(2026, 3, 5), "BURGER  KING 09812", 11.40),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        assert list(positives.keys()) == ["BURGER KING"]
        assert len(positives["BURGER KING"]) == 3

    def test_display_brand_prefers_spaced_form(self):
        # 1 spaced row vs 5 unspaced — rule 1 says spaced still wins.
        rows = [
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGER KING #1", 1.0),
        ] + [
            (datetime(2026, 3, i + 2), datetime(2026, 3, i + 2), f"BURGERKING #{i}", 1.0 + i)
            for i in range(5)
        ]
        positives, _ = group_and_sort(load_transactions(_build_input(rows)))
        assert list(positives.keys()) == ["BURGER KING"]
        assert len(positives["BURGER KING"]) == 6

    def test_display_brand_uses_most_common_when_no_spaced_form(self):
        # All forms collapse to the same string post-normalization, so the
        # display label is just that string.
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGERKING #1", 1.0),
            (datetime(2026, 3, 2), datetime(2026, 3, 2), "BURGERKING #1", 2.0),
            (datetime(2026, 3, 3), datetime(2026, 3, 3), "BURGERKING #1", 3.0),
            (datetime(2026, 3, 4), datetime(2026, 3, 4), "BURGERKING #2", 4.0),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        assert list(positives.keys()) == ["BURGERKING"]

    def test_hyphen_vs_space_does_not_merge(self):
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "7-ELEVEN 33252", 7.78),
            (datetime(2026, 3, 5), datetime(2026, 3, 5), "7 ELEVEN 38097", 13.22),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        # Two distinct brand groups.
        assert set(positives.keys()) == {"7-ELEVEN", "7 ELEVEN"}

    def test_abbreviation_does_not_merge(self):
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BK #4421", 7.50),
            (datetime(2026, 3, 5), datetime(2026, 3, 5), "BURGER KING #4421", 9.85),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        assert set(positives.keys()) == {"BK", "BURGER KING"}

    def test_displayed_merchant_name_unchanged_by_merge(self):
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGERKING #2231", 8.95),
            (datetime(2026, 3, 8), datetime(2026, 3, 8), "BURGER KING 09812", 11.40),
        ])
        ws = _data_ws(process_merchant_charges(data))
        # Column C of the data rows must contain the original raw strings,
        # NOT the brand label.
        merchant_cells = [
            ws.cell(row=r, column=3).value
            for r in range(2, ws.max_row + 1)
            if isinstance(ws.cell(row=r, column=1).value, datetime)
        ]
        assert "BURGER KING 09812" in merchant_cells
        assert "BURGERKING #2231" in merchant_cells
        # And the brand label "BURGER KING" itself does NOT appear as a
        # data-row value (it's only used internally as a group key).
        assert "BURGER KING" not in merchant_cells

    def test_within_merged_group_sort_order_preserved(self):
        data = _build_input([
            # All same brand after merge; varied raw strings + dates.
            (datetime(2026, 3, 8), datetime(2026, 3, 8), "BURGER KING 09812", 11.40),
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGER KING #4421", 9.85),
            (datetime(2026, 3, 5), datetime(2026, 3, 5), "BURGERKING #2231", 8.95),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        rows = positives["BURGER KING"]
        # _merchant_sort_key strips non-alphanumeric → numeric portion drives
        # the order: 09812 < 2231 < 4421.
        assert [r["merchant"] for r in rows] == [
            "BURGER KING 09812", "BURGERKING #2231", "BURGER KING #4421",
        ]

    def test_stress_test_5_burger_subset_produces_two_groups(self):
        """Inline simulation of the BURGER subset of Stress Test 5.

        The real Stress_Test_5_Real_World_Chaos.xlsx is not present in this
        repo, so this test exercises the same logical scenario: BK + three
        BURGER KING/BURGERKING variants must collapse to two brand groups
        (BK separate; the three burger-king-spelling variants merged into
        one group totaling $30.20).
        """
        data = _build_input([
            (datetime(2026, 3, 1), datetime(2026, 3, 1), "BURGER KING #4421", 9.85),
            (datetime(2026, 3, 5), datetime(2026, 3, 5), "BURGER KING 09812", 11.40),
            (datetime(2026, 3, 8), datetime(2026, 3, 8), "BURGERKING #2231", 8.95),
            (datetime(2026, 3, 10), datetime(2026, 3, 10), "BK #4421", 7.50),
            # plus a couple unrelated brands so this exercises the full pipeline
            (datetime(2026, 3, 12), datetime(2026, 3, 12), "STARBUCKS #1", 5.25),
            (datetime(2026, 3, 13), datetime(2026, 3, 13), "ARCO 912987", 47.97),
        ])
        positives, _ = group_and_sort(load_transactions(data))
        assert "BURGER KING" in positives
        assert "BK" in positives
        assert len(positives["BURGER KING"]) == 3
        assert abs(sum(r["amount"] for r in positives["BURGER KING"]) - 30.20) < 0.01
        assert len(positives["BK"]) == 1
        assert positives["BK"][0]["amount"] == 7.50


# ---------------------------------------------------------------------------
# end-to-end on the local fixture
# ---------------------------------------------------------------------------

class TestEndToEndFixture:
    def test_pipeline_with_sample_fixture(self):
        out = process_merchant_charges(FIXTURE_PATH.read_bytes())
        wb = openpyxl.load_workbook(io.BytesIO(out))
        assert wb.sheetnames == [SUMMARY_SHEET_NAME, OUTPUT_SHEET_NAME]
        assert wb.active.title == SUMMARY_SHEET_NAME

        positives, negatives = group_and_sort(load_transactions(FIXTURE_PATH.read_bytes()))
        assert set(positives.keys()) == {"ARCO", "7-ELEVEN", "76 - TEMECULA FUELS CORP", "THE HOME DEPOT"}
        assert len(negatives) == 1

        data_ws = wb[OUTPUT_SHEET_NAME]
        gts = _find_grand_total_rows(data_ws)
        assert "Grand Total (Charges Only)" in gts
        assert "Grand Total (Net)" in gts


# ---------------------------------------------------------------------------
# end-to-end on the real customer file
# ---------------------------------------------------------------------------

@pytest.mark.skipif(not CUSTOMER_INPUT.exists(), reason="customer Input.xlsx not on this host")
class TestEndToEndCustomerFile:
    def test_customer_file_layout(self):
        positives, negatives = group_and_sort(load_transactions(CUSTOMER_INPUT.read_bytes()))
        assert len(positives) == 26
        assert len(negatives) == 1
        assert negatives[0]["merchant"] == "PAYMENT RECEIVED - THANK"
        assert abs(negatives[0]["amount"] - (-5193.86)) < 0.001

    def test_customer_file_summary_and_grand_totals(self):
        out = process_merchant_charges(CUSTOMER_INPUT.read_bytes())
        wb = openpyxl.load_workbook(io.BytesIO(out), data_only=False)
        assert wb.sheetnames == [SUMMARY_SHEET_NAME, OUTPUT_SHEET_NAME]
        assert wb.active.title == SUMMARY_SHEET_NAME

        sum_ws = wb[SUMMARY_SHEET_NAME]
        assert _summary_value_for(sum_ws, "Total transactions in input:") == 54
        assert _summary_value_for(sum_ws, "Charges (positive):") == 53
        assert _summary_value_for(sum_ws, "Credits (negative):") == 1
        assert not _summary_has_label(sum_ws, "Rows dropped (validation):")
        assert _summary_value_for(sum_ws, "Total brand groups:") == 26

        # Resolve the Charges Only formula to its computed value.
        data_ws = wb[OUTPUT_SHEET_NAME]
        charges_formula = _summary_value_for(sum_ws, "Grand Total (Charges Only):")
        ref_row = int(charges_formula.split("!E")[1])
        charges = _eval_brand_subtotals(data_ws, data_ws.cell(row=ref_row, column=5).value)
        assert abs(charges - 5046.98) < 0.01

        # Net = Charges + Credits subtotal (referenced via the data sheet's E_charges + E_credits formula).
        net_formula = _summary_value_for(sum_ws, "Grand Total (Net):")
        ref_row = int(net_formula.split("!E")[1])
        # Data-sheet net formula is e.g. "=E111+E109".
        net_cell_formula = data_ws.cell(row=ref_row, column=5).value
        a, b = net_cell_formula.lstrip("=").split("+")
        a_row = int(a.strip()[1:])
        b_row = int(b.strip()[1:])
        a_val = _eval_brand_subtotals(data_ws, data_ws.cell(row=a_row, column=5).value)
        b_val = _eval_sum_range(data_ws, data_ws.cell(row=b_row, column=4).value)
        assert abs((a_val + b_val) - (-146.88)) < 0.01

    def test_customer_data_sheet_layout_unchanged(self):
        # Regression guard: data sheet starts with a real transaction row, no inline summary block.
        out = process_merchant_charges(CUSTOMER_INPUT.read_bytes())
        ws = openpyxl.load_workbook(io.BytesIO(out), data_only=False)[OUTPUT_SHEET_NAME]
        assert [c.value for c in ws[1]] == list(
            ("Posted Date", "Occurred Date", "Merchant Name", "Amount", "Merchant Total")
        )
        assert isinstance(ws.cell(row=2, column=1).value, datetime)
        assert ws.freeze_panes == "A2"
