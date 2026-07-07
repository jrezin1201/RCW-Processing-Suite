"""Tests for the Capital One Card Report module.

Covers loader validation (required columns, blank Category/Debit/Credit),
grouping + sort order, transaction-level detail preservation (no merging),
the data sheet layout (headers, bold category rows + subtotals, blank amount
cells, single Grand Total), the Analytics sheet, and the real customer
sample CSV.
"""
from __future__ import annotations

import io
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from app.modules.capital_one_card.services import (
    ANALYTICS_SHEET_NAME,
    GRAND_TOTAL_LABEL,
    OUTPUT_HEADERS,
    OUTPUT_SHEET_NAME,
    UNCATEGORIZED_LABEL,
    CapitalOneCardError,
    group_and_sort,
    load_transactions,
    process_capital_one_card,
    write_report,
)

FIXTURE_PATH = Path(__file__).parent / "fixtures" / "capital_one_card_sample.csv"
CUSTOMER_INPUT = Path(
    "/Users/jordanhill/Desktop/RCwendt-Colissa/05-CapitalOne-Card/Trev June 2026.csv"
)

FULL_HEADER = "Transaction Date,Posted Date,Card No.,Description,Category,Debit,Credit"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

def _build_csv(rows: list[str], header: str = FULL_HEADER) -> bytes:
    return ("\n".join([header, *rows]) + "\n").encode("utf-8")


def _build_xlsx(rows: list[list], header: str = FULL_HEADER) -> bytes:
    """Build an xlsx input file; rows hold native cell values (datetime, float, ...)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header.split(","))
    for row in rows:
        ws.append(row)
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _sheet(wb_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)[OUTPUT_SHEET_NAME]


def _analytics(wb_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(wb_bytes), data_only=False)[ANALYTICS_SHEET_NAME]


def _analytics_value(ws, label: str):
    """Find a 'Label:' row in the Analytics key/value area and return its value."""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=2).value == label:
            return ws.cell(row=r, column=3).value
    pytest.fail(f"Analytics label not found: {label}")


def _grand_total_row(ws) -> int:
    rows = [
        r for r in range(2, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == GRAND_TOTAL_LABEL
    ]
    assert len(rows) == 1, f"expected exactly one Grand Total row, found {len(rows)}"
    return rows[0]


def _eval_sum(ws, formula: str) -> float:
    """Evaluate a literal '=SUM(C{a}:C{b})' against the written cell values."""
    assert formula.startswith("=SUM(") and formula.endswith(")")
    start, end = formula[len("=SUM("):-1].split(":")
    col_idx = openpyxl.utils.column_index_from_string(start[0])
    total = 0.0
    for r in range(int(start[1:]), int(end[1:]) + 1):
        v = ws.cell(row=r, column=col_idx).value
        if isinstance(v, int | float):
            total += float(v)
    return total


def _eval_grand_total(ws, formula: str) -> float:
    """Evaluate '=SUM(C{a},C{b},...)' by walking each referenced subtotal cell
    down to its underlying =SUM(C{x}:C{y}) range over the raw amounts."""
    assert formula.startswith("=SUM(") and formula.endswith(")")
    total = 0.0
    for ref in formula[len("=SUM("):-1].split(","):
        ref = ref.strip()
        col_idx = openpyxl.utils.column_index_from_string(ref[0])
        total += _eval_sum(ws, ws.cell(row=int(ref[1:]), column=col_idx).value)
    return total


def _subtotal_rows(ws) -> list[int]:
    """Rows holding description-block subtotals: a =SUM formula in C or D with
    no date/description, excluding the Grand Total row."""
    gt_row = _grand_total_row(ws)
    out = []
    for r in range(2, ws.max_row + 1):
        if r == gt_row:
            continue
        c_val = ws.cell(row=r, column=3).value
        d_val = ws.cell(row=r, column=4).value
        if isinstance(c_val, str) and c_val.startswith("=SUM(") or \
           isinstance(d_val, str) and d_val.startswith("=SUM("):
            out.append(r)
    return out


# ---------------------------------------------------------------------------
# loader
# ---------------------------------------------------------------------------

class TestLoader:
    def test_loads_all_rows(self):
        data = _build_csv([
            "6/15/2026,6/16/2026,8102,VENDOR A,Other,85,",
            "6/16/2026,6/16/2026,8102,VENDOR B,Merchandise,410,",
        ])
        txs = load_transactions(data)
        assert len(txs) == 2
        assert txs[0]["description"] == "VENDOR A"
        assert txs[0]["debit"] == 85.0
        assert txs[0]["credit"] is None
        assert txs[0]["date"] == datetime(2026, 6, 15)

    @pytest.mark.parametrize("missing", ["Transaction Date", "Description", "Category", "Debit", "Credit"])
    def test_missing_required_column_is_named_in_error(self, missing):
        headers = [h for h in FULL_HEADER.split(",") if h != missing]
        data = _build_csv(["x," * (len(headers) - 1) + "x"], header=",".join(headers))
        with pytest.raises(CapitalOneCardError, match=missing):
            load_transactions(data)

    def test_posted_date_and_card_no_are_optional(self):
        data = _build_csv(
            ["6/15/2026,VENDOR A,Other,85,"],
            header="Transaction Date,Description,Category,Debit,Credit",
        )
        assert len(load_transactions(data)) == 1

    def test_blank_category_becomes_uncategorized(self):
        data = _build_csv(["6/15/2026,6/16/2026,8102,VENDOR A,,85,"])
        assert load_transactions(data)[0]["category"] == UNCATEGORIZED_LABEL

    def test_blank_debit_and_credit_stay_none(self):
        data = _build_csv(["6/15/2026,6/16/2026,8102,VENDOR A,Other,,"])
        tx = load_transactions(data)[0]
        assert tx["debit"] is None
        assert tx["credit"] is None

    def test_amounts_with_dollar_signs_and_commas_parse(self):
        data = _build_csv(['6/15/2026,6/16/2026,8102,VENDOR A,Other,"$1,234.56",'])
        assert load_transactions(data)[0]["debit"] == 1234.56

    def test_quoted_description_with_comma(self):
        data = _build_csv(['5/31/2026,6/1/2026,8102,"STERICYCLE, INC",Other Services,120,'])
        assert load_transactions(data)[0]["description"] == "STERICYCLE, INC"

    def test_utf8_bom_is_tolerated(self):
        data = b"\xef\xbb\xbf" + _build_csv(["6/15/2026,6/16/2026,8102,VENDOR A,Other,85,"])
        assert len(load_transactions(data)) == 1

    def test_empty_file_raises(self):
        with pytest.raises(CapitalOneCardError, match="empty"):
            load_transactions(b"")

    def test_header_only_raises(self):
        with pytest.raises(CapitalOneCardError, match="No transactions"):
            load_transactions((FULL_HEADER + "\n").encode())

    def test_unparseable_date_passes_through_as_string(self):
        data = _build_csv(["June 15,6/16/2026,8102,VENDOR A,Other,85,"])
        # 'June 15' splits the quoted field? No quotes — 'June 15' has no comma.
        assert load_transactions(data)[0]["date"] == "June 15"


# ---------------------------------------------------------------------------
# xlsx loader
# ---------------------------------------------------------------------------

class TestXlsxLoader:
    SAMPLE = [
        [datetime(2026, 6, 15), datetime(2026, 6, 16), 8102, "VENDOR A", "Other", 85.0, None],
        [datetime(2026, 6, 16), datetime(2026, 6, 16), 8102, "VENDOR B", "Merchandise", 410.0, None],
    ]

    def test_loads_native_excel_types(self):
        txs = load_transactions(_build_xlsx(self.SAMPLE))
        assert len(txs) == 2
        assert txs[0]["date"] == datetime(2026, 6, 15)
        assert txs[0]["description"] == "VENDOR A"
        assert txs[0]["debit"] == 85.0
        assert txs[0]["credit"] is None

    def test_numeric_card_number_reads_without_decimal(self):
        txs = load_transactions(_build_xlsx(self.SAMPLE))
        assert txs[0]["card"] == "8102"

    def test_text_dates_and_amounts_also_parse(self):
        txs = load_transactions(_build_xlsx([
            ["6/15/2026", "6/16/2026", "8102", "VENDOR A", "Other", "$1,234.56", ""],
        ]))
        assert txs[0]["date"] == datetime(2026, 6, 15)
        assert txs[0]["debit"] == 1234.56

    def test_blank_category_becomes_uncategorized(self):
        txs = load_transactions(_build_xlsx([
            [datetime(2026, 6, 15), None, 8102, "VENDOR A", None, 85.0, None],
        ]))
        assert txs[0]["category"] == UNCATEGORIZED_LABEL

    def test_missing_required_column_raises(self):
        data = _build_xlsx(
            [[datetime(2026, 6, 15), "VENDOR A", 85.0, None]],
            header="Transaction Date,Description,Debit,Credit",
        )
        with pytest.raises(CapitalOneCardError, match="Category"):
            load_transactions(data)

    def test_header_only_raises(self):
        with pytest.raises(CapitalOneCardError, match="No transactions"):
            load_transactions(_build_xlsx([]))

    def test_all_blank_rows_are_skipped(self):
        txs = load_transactions(_build_xlsx([
            [None, None, None, None, None, None, None],
            [datetime(2026, 6, 15), None, 8102, "VENDOR A", "Other", 85.0, None],
        ]))
        assert len(txs) == 1

    def test_xlsx_and_csv_inputs_produce_identical_reports(self):
        csv_rows = [
            "6/15/2026,6/16/2026,8102,DUNN-EDWARDS VT CORP 9,Merchandise,81.48,",
            "6/14/2026,6/15/2026,8102,CROSSFIT SURF CITY,Entertainment,4,",
            "6/2/2026,6/4/2026,8102,THE HOME DEPOT #6963,Merchandise,,125.58",
        ]
        xlsx_rows = [
            [datetime(2026, 6, 15), datetime(2026, 6, 16), 8102, "DUNN-EDWARDS VT CORP 9", "Merchandise", 81.48, None],
            [datetime(2026, 6, 14), datetime(2026, 6, 15), 8102, "CROSSFIT SURF CITY", "Entertainment", 4, None],
            [datetime(2026, 6, 2), datetime(2026, 6, 4), 8102, "THE HOME DEPOT #6963", "Merchandise", None, 125.58],
        ]
        from_csv = load_transactions(_build_csv(csv_rows))
        from_xlsx = load_transactions(_build_xlsx(xlsx_rows))
        assert from_csv == from_xlsx

    def test_xlsx_processes_end_to_end(self):
        out = process_capital_one_card(_build_xlsx(self.SAMPLE))
        ws = _sheet(out)
        gt_row = _grand_total_row(ws)
        assert _eval_grand_total(ws, ws.cell(row=gt_row, column=3).value) == pytest.approx(495.0)


# ---------------------------------------------------------------------------
# grouping + sorting
# ---------------------------------------------------------------------------

class TestGroupAndSort:
    def test_categories_sorted_ascending(self):
        data = _build_csv([
            "6/1/2026,6/2/2026,8102,V1,Merchandise,10,",
            "6/1/2026,6/2/2026,8102,V2,Dining,20,",
            "6/1/2026,6/2/2026,8102,V3,Entertainment,30,",
        ])
        grouped = group_and_sort(load_transactions(data))
        assert list(grouped.keys()) == ["Dining", "Entertainment", "Merchandise"]

    def test_descriptions_sorted_within_category(self):
        data = _build_csv([
            "6/1/2026,6/2/2026,8102,ZEBRA,Merchandise,10,",
            "6/1/2026,6/2/2026,8102,apple,Merchandise,20,",
            "6/1/2026,6/2/2026,8102,MANGO,Merchandise,30,",
        ])
        grouped = group_and_sort(load_transactions(data))
        descs = [t["description"] for t in grouped["Merchandise"]]
        assert descs == ["apple", "MANGO", "ZEBRA"]

    def test_same_description_sorted_by_date(self):
        data = _build_csv([
            "6/10/2026,6/11/2026,8102,VENDOR,Merchandise,10,",
            "6/1/2026,6/2/2026,8102,VENDOR,Merchandise,20,",
            "6/5/2026,6/6/2026,8102,VENDOR,Merchandise,30,",
        ])
        grouped = group_and_sort(load_transactions(data))
        dates = [t["date"] for t in grouped["Merchandise"]]
        assert dates == sorted(dates)

    def test_duplicate_descriptions_never_merged(self):
        data = _build_csv([
            "6/1/2026,6/2/2026,8102,THEFARMERSDOG.COM,Merchandise,410,",
            "6/8/2026,6/9/2026,8102,THEFARMERSDOG.COM,Merchandise,410,",
            "6/15/2026,6/16/2026,8102,THEFARMERSDOG.COM,Merchandise,410,",
        ])
        grouped = group_and_sort(load_transactions(data))
        assert len(grouped["Merchandise"]) == 3

    def test_no_rows_lost_in_grouping(self):
        txs = load_transactions(FIXTURE_PATH.read_bytes())
        grouped = group_and_sort(txs)
        assert sum(len(rows) for rows in grouped.values()) == len(txs)


# ---------------------------------------------------------------------------
# output sheet
# ---------------------------------------------------------------------------

class TestOutputSheet:
    SAMPLE = [
        "6/15/2026,6/16/2026,8102,DUNN-EDWARDS VT CORP 9,Merchandise,81.48,",
        "6/16/2026,6/16/2026,8102,THEFARMERSDOG.COM,Merchandise,410,",
        "6/14/2026,6/15/2026,8102,CROSSFIT SURF CITY,Entertainment,4,",
        "6/2/2026,6/4/2026,8102,THE HOME DEPOT #6963,Merchandise,,125.58",
    ]

    def _out(self):
        return process_capital_one_card(_build_csv(self.SAMPLE))

    def test_header_row_bold_and_frozen(self):
        ws = _sheet(self._out())
        assert ws.freeze_panes == "A2"
        for col, label in enumerate(OUTPUT_HEADERS, start=1):
            cell = ws.cell(row=1, column=col)
            assert cell.value == label
            assert cell.font.bold

    def test_only_four_output_columns(self):
        ws = _sheet(self._out())
        assert ws.max_column == len(OUTPUT_HEADERS)
        # Posted Date / Card No. values must not leak anywhere in the sheet.
        for row in ws.iter_rows(values_only=True):
            assert "8102" not in [str(v) for v in row if v is not None]

    def test_category_rows_bold(self):
        ws = _sheet(self._out())
        cat_rows = [
            r for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=1).value in ("Entertainment", "Merchandise")
        ]
        assert len(cat_rows) == 2
        for r in cat_rows:
            assert ws.cell(row=r, column=1).font.bold

    def test_category_order_and_rows_within(self):
        ws = _sheet(self._out())
        col_a = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        ent_idx = col_a.index("Entertainment")
        mer_idx = col_a.index("Merchandise")
        assert ent_idx < mer_idx
        descs = [
            ws.cell(row=r, column=2).value
            for r in range(mer_idx + 3, ws.max_row + 1)
            if ws.cell(row=r, column=2).value
        ]
        assert descs == sorted(descs, key=str.casefold)

    def test_blank_amount_cells_stay_blank(self):
        ws = _sheet(self._out())
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "THE HOME DEPOT #6963":
                assert ws.cell(row=r, column=3).value is None
                assert ws.cell(row=r, column=4).value == 125.58
                break
        else:
            pytest.fail("THE HOME DEPOT #6963 row not found")

    def test_transaction_dates_preserved(self):
        ws = _sheet(self._out())
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=2).value == "DUNN-EDWARDS VT CORP 9":
                assert ws.cell(row=r, column=1).value == datetime(2026, 6, 15)
                break
        else:
            pytest.fail("DUNN-EDWARDS row not found")

    def test_single_grand_total_at_bottom_bold_with_correct_sums(self):
        ws = _sheet(self._out())
        gt_row = _grand_total_row(ws)
        assert gt_row == ws.max_row
        assert ws.cell(row=gt_row, column=1).font.bold
        debit_formula = ws.cell(row=gt_row, column=3).value
        credit_formula = ws.cell(row=gt_row, column=4).value
        assert _eval_grand_total(ws, debit_formula) == pytest.approx(81.48 + 410 + 4)
        assert _eval_grand_total(ws, credit_formula) == pytest.approx(125.58)

    def test_each_category_block_has_bold_labeled_subtotal(self):
        ws = _sheet(self._out())
        # 2 categories in the sample -> exactly 2 subtotal rows.
        subs = _subtotal_rows(ws)
        assert len(subs) == 2
        labels = [ws.cell(row=r, column=1).value for r in subs]
        assert labels == ["Entertainment Total", "Merchandise Total"]
        for r in subs:
            assert ws.cell(row=r, column=1).font.bold
            assert ws.cell(row=r, column=3).font.bold

    def test_category_subtotal_sums_the_whole_category(self):
        ws = _sheet(self._out())
        for r in _subtotal_rows(ws):
            if ws.cell(row=r, column=1).value == "Merchandise Total":
                assert _eval_sum(ws, ws.cell(row=r, column=3).value) == pytest.approx(81.48 + 410)
                assert _eval_sum(ws, ws.cell(row=r, column=4).value) == pytest.approx(125.58)
                break
        else:
            pytest.fail("Merchandise Total row not found")

    def test_grand_total_counts_each_transaction_exactly_once(self):
        ws = _sheet(self._out())
        gt_row = _grand_total_row(ws)
        raw_debits = sum(
            float(ws.cell(row=r, column=3).value)
            for r in range(2, gt_row)
            if r not in _subtotal_rows(ws)
            and isinstance(ws.cell(row=r, column=3).value, int | float)
        )
        assert _eval_grand_total(ws, ws.cell(row=gt_row, column=3).value) == pytest.approx(raw_debits)

    def test_amount_cells_two_decimal_format(self):
        ws = _sheet(self._out())
        for r in range(2, ws.max_row + 1):
            for c in (3, 4):
                cell = ws.cell(row=r, column=c)
                if isinstance(cell.value, int | float):
                    assert "0.00" in cell.number_format

    def test_write_report_empty_group_still_produces_grand_total(self):
        out = write_report(group_and_sort(load_transactions(_build_csv(self.SAMPLE[:1]))))
        ws = _sheet(out)
        _grand_total_row(ws)


# ---------------------------------------------------------------------------
# analytics sheet
# ---------------------------------------------------------------------------

class TestAnalyticsSheet:
    SAMPLE = TestOutputSheet.SAMPLE

    def _out(self):
        return process_capital_one_card(_build_csv(self.SAMPLE))

    def test_analytics_sheet_first_and_active(self):
        wb = openpyxl.load_workbook(io.BytesIO(self._out()))
        assert wb.sheetnames[0] == ANALYTICS_SHEET_NAME
        assert wb.active.title == ANALYTICS_SHEET_NAME

    def test_total_transaction_count(self):
        ws = _analytics(self._out())
        assert _analytics_value(ws, "Total transactions:") == 4

    def test_card_numbers_listed(self):
        ws = _analytics(self._out())
        assert _analytics_value(ws, "Card number(s):") == "8102"

    def test_multiple_card_numbers_joined(self):
        rows = self.SAMPLE + ["6/1/2026,6/2/2026,9944,VENDOR X,Dining,10,"]
        ws = _analytics(process_capital_one_card(_build_csv(rows)))
        assert _analytics_value(ws, "Card number(s):") == "8102, 9944"

    def test_missing_card_column_noted(self):
        data = _build_csv(
            ["6/15/2026,VENDOR A,Other,85,"],
            header="Transaction Date,Description,Category,Debit,Credit",
        )
        ws = _analytics(process_capital_one_card(data))
        assert _analytics_value(ws, "Card number(s):") == "Not in input file"

    def test_grand_totals_reference_data_sheet(self):
        out = self._out()
        data_ws = _sheet(out)
        gt_row = _grand_total_row(data_ws)
        ws = _analytics(out)
        assert _analytics_value(ws, "Grand Total (Debit):") == f"='{OUTPUT_SHEET_NAME}'!C{gt_row}"
        assert _analytics_value(ws, "Grand Total (Credit):") == f"='{OUTPUT_SHEET_NAME}'!D{gt_row}"

    def test_by_category_counts_and_totals(self):
        out = self._out()
        data_ws = _sheet(out)
        sub_rows = {data_ws.cell(row=r, column=1).value: r for r in _subtotal_rows(data_ws)}
        ws = _analytics(out)
        table = {}
        for r in range(1, ws.max_row + 1):
            name = ws.cell(row=r, column=2).value
            if name in ("Entertainment", "Merchandise"):
                table[name] = (
                    ws.cell(row=r, column=3).value,
                    ws.cell(row=r, column=4).value,
                    ws.cell(row=r, column=5).value,
                )
        assert table["Entertainment"][0] == 1
        assert table["Merchandise"][0] == 3
        for name in ("Entertainment", "Merchandise"):
            sub_row = sub_rows[f"{name} Total"]
            assert table[name][1] == f"='{OUTPUT_SHEET_NAME}'!C{sub_row}"
            assert table[name][2] == f"='{OUTPUT_SHEET_NAME}'!D{sub_row}"


# ---------------------------------------------------------------------------
# customer sample file
# ---------------------------------------------------------------------------

class TestSampleFile:
    def test_fixture_processes_end_to_end(self):
        txs = load_transactions(FIXTURE_PATH.read_bytes())
        assert len(txs) == 92
        out = process_capital_one_card(FIXTURE_PATH.read_bytes())
        ws = _sheet(out)
        gt_row = _grand_total_row(ws)
        assert _eval_grand_total(ws, ws.cell(row=gt_row, column=3).value) == pytest.approx(233737.47, abs=0.01)
        assert _eval_grand_total(ws, ws.cell(row=gt_row, column=4).value) == pytest.approx(208193.00, abs=0.01)

    def test_fixture_transaction_rows_all_present(self):
        txs = load_transactions(FIXTURE_PATH.read_bytes())
        out = process_capital_one_card(FIXTURE_PATH.read_bytes())
        ws = _sheet(out)
        data_rows = [
            r for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=2).value  # description present => transaction row
        ]
        assert len(data_rows) == len(txs)

    @pytest.mark.skipif(not CUSTOMER_INPUT.exists(), reason="customer CSV not on this host")
    def test_customer_file_processes(self):
        out = process_capital_one_card(CUSTOMER_INPUT.read_bytes())
        ws = _sheet(out)
        _grand_total_row(ws)
