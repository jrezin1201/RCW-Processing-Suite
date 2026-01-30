#!/usr/bin/env python3
"""
Script to extract project metadata from a spreadsheet.
Extracts project name from B3 and phase number from B5.
"""

import sys
import re
import openpyxl
from pathlib import Path


def extract_project_metadata(file_path):
    """
    Extract project name and phase number from specified cells in an Excel file.

    Args:
        file_path (str): Path to the Excel file

    Returns:
        tuple: (project_name, phase_number)
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # 1. Extract Project Name from B3
        b3_value = ws['B3'].value
        project_name = None

        if b3_value:
            # Convert to string and split by hyphens
            parts = str(b3_value).split('-')

            # Extract word between first and second hyphen
            if len(parts) >= 2:
                # The second part (index 1) should contain the project name
                project_name = parts[1].strip()
            else:
                print(f"Warning: Could not extract project name from B3: '{b3_value}'")
        else:
            print("Warning: Cell B3 is empty")

        # 2. Extract Phase Number from B5
        b5_value = ws['B5'].value
        phase_number = None

        if b5_value:
            # Look for pattern PH followed by digits
            phase_match = re.search(r'PH(\d{2})', str(b5_value), re.IGNORECASE)

            if phase_match:
                # Extract the digits and convert to integer (removes leading zeros)
                phase_digits = phase_match.group(1)
                phase_number = int(phase_digits)
            else:
                print(f"Warning: Could not find phase pattern in B5: '{b5_value}'")
        else:
            print("Warning: Cell B5 is empty")

        # Close the workbook
        wb.close()

        return project_name, phase_number

    except FileNotFoundError:
        print(f"Error: File not found: {file_path}")
        return None, None
    except Exception as e:
        print(f"Error reading file: {e}")
        return None, None


def main():
    """Main function to run the metadata extraction."""

    # Check if file path is provided as argument
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # If no argument provided, prompt for file path
        file_path = input("Enter the path to the Excel file: ").strip()

        # Remove quotes if present
        if file_path.startswith('"') and file_path.endswith('"'):
            file_path = file_path[1:-1]

    # Check if file exists
    if not Path(file_path).exists():
        print(f"Error: File does not exist: {file_path}")
        sys.exit(1)

    # Extract metadata
    project_name, phase_number = extract_project_metadata(file_path)

    # Output results
    if project_name or phase_number:
        project_str = project_name if project_name else "[Not Found]"
        phase_str = str(phase_number) if phase_number is not None else "[Not Found]"
        print(f"\nProject: {project_str} | Phase: {phase_str}")
    else:
        print("\nError: Could not extract metadata from the file")
        sys.exit(1)


if __name__ == "__main__":
    main()