#!/usr/bin/env python3
"""
Debug script to analyze an Excel file and show what the app sees.
This helps diagnose why "Location Total" rows might not be found.
"""

import sys
import openpyxl
import re

def debug_excel_file(filepath):
    """Analyze Excel file to help debug processing issues."""

    print("="*60)
    print(f"ANALYZING FILE: {filepath}")
    print("="*60)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        print(f"\n📊 Sheet Name: {ws.title}")
        print(f"📊 Total Rows: {ws.max_row}")
        print(f"📊 Total Columns: {ws.max_column}")

        # Show first 3 rows to understand structure
        print("\n🔍 FIRST 3 ROWS (to identify headers):")
        print("-"*60)
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num > 3:
                break
            print(f"Row {row_num}:")
            for col_idx, value in enumerate(row[:12] if row else []):  # Show first 12 columns
                col_letter = chr(65 + col_idx)  # A, B, C, etc.
                print(f"  Col {col_letter} (idx {col_idx}): {repr(value)}")
            print()

        # Look for "Location Total" anywhere
        print("\n🔍 SEARCHING FOR 'Location Total' ROWS:")
        print("-"*60)
        location_total_found = False

        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row and len(row) >= 4:
                # Check column D (index 3) for "Location Total"
                employee_col = row[3] if len(row) > 3 else None

                if employee_col and "location total" in str(employee_col).lower():
                    location_total_found = True
                    location_col = row[1] if len(row) > 1 else None
                    hours_col = row[11] if len(row) > 11 else None

                    print(f"✓ Found at Row {row_num}:")
                    print(f"  Location (Col B): {repr(location_col)}")
                    print(f"  Employee (Col D): {repr(employee_col)}")
                    print(f"  Total Hours (Col L): {repr(hours_col)}")

                    # Check if job number can be extracted
                    if location_col:
                        match = re.match(r'^\s*(\d{4})\b', str(location_col))
                        if match:
                            print(f"  ✓ Job Number Found: {match.group(1)}")
                        else:
                            print(f"  ✗ No 4-digit job number at start of location")
                    print()

        if not location_total_found:
            print("❌ NO 'Location Total' found in Column D!")
            print("\n🔍 CHECKING ALL COLUMN D VALUES (first 20 rows):")
            for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_num > 20:
                    break
                if row and len(row) > 3:
                    print(f"  Row {row_num}: {repr(row[3])}")

        # Check if columns might be shifted
        print("\n🔍 LOOKING FOR 'Location Total' IN ANY COLUMN:")
        print("-"*60)
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num > 10:  # Check first 10 rows
                break
            if row:
                for col_idx, value in enumerate(row):
                    if value and "location total" in str(value).lower():
                        col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                        print(f"Found 'Location Total' at Row {row_num}, Column {col_letter} (index {col_idx}): {repr(value)}")

        # Show what app expects
        print("\n📋 WHAT THE APP EXPECTS:")
        print("-"*60)
        print("• Column B (index 1): Location with job number starting with 4 digits")
        print("• Column D (index 3): Employee Name, looking for 'Location Total'")
        print("• Column L (index 11): Total Hours (numeric value)")
        print("\n⚠️  Note: Columns are 0-indexed in code, so B=1, D=3, L=11")

    except Exception as e:
        print(f"\n❌ ERROR reading file: {e}")
        print("\nMake sure the file is a valid .xlsx file")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        debug_excel_file(sys.argv[1])
    else:
        print("Usage: python debug_excel.py your_file.xlsx")
        print("\nThis script will analyze your Excel file to help debug why")
        print("'Location Total' rows might not be found.")