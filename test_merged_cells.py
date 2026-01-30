#!/usr/bin/env python3
"""Test that cells B1 and C1 are merged for Project Name."""
import openpyxl
from openpyxl import Workbook

# Create test data with metadata
wb = Workbook()
ws = wb.active

# Add metadata in specific cells
ws['B3'] = "OCH - TestProject - 123456"  # Project name: "TestProject"
ws['B5'] = "PH09 - HS 0100-0105"  # Phase and House String

# Add test data
test_data = [
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["Lot/Block", "Plan", "Elevation", "Task", "Task Start Date", "Subtotal"],
    ["44", "2", "B", "Prime Exterior", "2024-01-01", "100.00"],
    ["45", "3", "A", "Paint Exterior", "2024-01-02", "150.00"],
]

for row_idx, row_data in enumerate(test_data[6:], 7):  # Start from row 7
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

test_filename = "test_merged_input.xlsx"
wb.save(test_filename)
wb.close()

print(f"Created test input file: {test_filename}")

# Process the file
from app.services.parser_lennar import parse_lennar_export
from app.services.aggregator import aggregate_data
from app.services.excel_writer import write_summary_excel

parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(test_filename)
summary_rows, qa_report = aggregate_data(parsed_rows, qa_meta)
output_path = write_summary_excel(
    summary_rows, qa_report, "test_merged",
    phase=phase, project_name=project_name, house_string=house_string
)

print(f"Output written to: {output_path}")

# Verify the merged cells
wb = openpyxl.load_workbook(output_path)
ws = wb.active

print("\n=== Merged Cells Check ===")

# Check if B1:C1 is in the merged cells ranges
merged_ranges = ws.merged_cells.ranges
print(f"Merged cell ranges in worksheet: {merged_ranges}")

# Check if B1:C1 is merged
b1_c1_merged = False
for merged_range in merged_ranges:
    if 'B1:C1' in str(merged_range):
        b1_c1_merged = True
        print(f"✅ B1:C1 is merged: {merged_range}")
        break

if not b1_c1_merged:
    # Check alternative format
    for merged_range in merged_ranges:
        min_row = merged_range.min_row
        max_row = merged_range.max_row
        min_col = merged_range.min_col
        max_col = merged_range.max_col

        # B1 is row 1, column 2; C1 is row 1, column 3
        if min_row == 1 and max_row == 1 and min_col == 2 and max_col == 3:
            b1_c1_merged = True
            print(f"✅ B1:C1 is merged (detected by coordinates): {merged_range}")
            break

if not b1_c1_merged:
    print("❌ B1:C1 is NOT merged")

# Check the content of the merged cell
b1_value = ws["B1"].value
c1_value = ws["C1"].value

print(f"\n=== Cell Contents ===")
print(f"B1 value: '{b1_value}'")
print(f"C1 value: '{c1_value}' (should be None/empty when merged)")

# Check that project name is displayed correctly
if b1_value and "Project Name:" in str(b1_value):
    print(f"✅ Project Name is correctly displayed in merged cell")
else:
    print(f"❌ Project Name not found in B1")

# Check other title row cells
print(f"\n=== Other Title Row Cells ===")
print(f"G1 (Phase): '{ws['G1'].value}'")
print(f"H1 (House String): '{ws['H1'].value}'")
print(f"I1 (Job#): '{ws['I1'].value}'")

# Check that headers in row 3 are not affected
print(f"\n=== Header Row Check (Row 3) ===")
print(f"B3 (LOT): '{ws['B3'].value}'")
print(f"C3 (PLAN): '{ws['C3'].value}'")

if b1_c1_merged and b1_value and "Project Name:" in str(b1_value):
    print("\n✅ SUCCESS: Cells B1 and C1 are properly merged with Project Name!")
    print("The Project Name field now spans across both columns B and C.")
else:
    print("\n⚠️ Some issues detected with cell merging")

wb.close()