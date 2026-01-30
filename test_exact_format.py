#!/usr/bin/env python3
"""Test that Excel output matches exact formatting requirements."""
import openpyxl
from openpyxl import Workbook

# Create test data
wb = Workbook()
ws = wb.active

ws['B2'] = "OCH - TestProject - 123456"  # Row 2 for some files
ws['B3'] = "OCH - TestProject - 123456"  # Row 3 for other files
ws['B4'] = "PH07 - HS 0044-0049, 0144-0142"
ws['B5'] = "PH07 - HS 0044-0049, 0144-0142"

# Add test data
test_data = [
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["Lot/Block", "Plan", "Elevation", "Task", "Task Start Date", "Subtotal"],
    ["44", "2B", "B", "Prime Exterior", "2024-01-01", "1543.50"],
    ["45", "3A", "A", "Paint Exterior", "2024-01-02", "1516.49"],
    ["46", "1B", "C", "Exterior UA", "2024-01-03", "1056.00"],
    ["47", "2C", "D", "Interior", "2024-01-04", "4115.99"],
]

for row_idx, row_data in enumerate(test_data[6:], 7):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

test_filename = "test_exact_format_input.xlsx"
wb.save(test_filename)
wb.close()

print("Created test input file")

# Process the file
from app.services.parser_lennar import parse_lennar_export
from app.services.aggregator import aggregate_data
from app.services.excel_writer import write_summary_excel

parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(test_filename)
summary_rows, qa_report = aggregate_data(parsed_rows, qa_meta)
output_path = write_summary_excel(
    summary_rows, qa_report, "test_exact_format",
    phase=phase, project_name=project_name, house_string=house_string
)

print(f"Output written to: {output_path}")

# Verify the exact formatting
wb = openpyxl.load_workbook(output_path)
ws = wb.active

print("\n" + "=" * 60)
print("EXACT FORMAT VERIFICATION")
print("=" * 60)

# Check Row 1
print("\nROW 1 (Project Header):")
print(f"  B1: '{ws['B1'].value}' (should have 'Project Name: ...')")
print(f"  G1: '{ws['G1'].value}' (should have 'Phase:X' with no space)")
print(f"  H1: '{ws['H1'].value}' (should have house string)")
print(f"  H1 has light blue fill: {ws['H1'].fill.start_color.index if ws['H1'].fill and ws['H1'].fill.start_color else 'No fill'}")
print(f"  I1: '{ws['I1'].value}' (should be 'Job #:')")

# Check Row 2
print("\nROW 2:")
print(f"  B2: '{ws['B2'].value}' (should be empty/None)")

# Check Row 3 (Headers)
print("\nROW 3 (Headers):")
print(f"  A3: '{ws['A3'].value}' (should be empty)")
print(f"  B3: '{ws['B3'].value}' (should be 'LOT')")
print(f"  C3: '{ws['C3'].value}' (should be 'PLAN')")
print(f"  D3: '{ws['D3'].value}' (should be 'EXT PRIME')")
print(f"  E3: '{ws['E3'].value}' (should be 'EXTERIOR')")
print(f"  K3: '{ws['K3'].value}' (should be 'Total')")
k3_fill = ws['K3'].fill.start_color.index if ws['K3'].fill and ws['K3'].fill.start_color else 'No fill'
print(f"  K3 fill color: {k3_fill} (should be FFFF00 for yellow)")

# Check data rows
print("\nDATA ROWS (starting at row 4):")
print(f"  A4: '{ws['A4'].value}' (should be '1' - row number)")
print(f"  B4: '{ws['B4'].value}' (should be LOT number)")
print(f"  K4 format: {ws['K4'].number_format} (should show currency format)")

# Find TOTAL row
total_found = False
total_row = 0
for row in range(4, 20):
    if ws[f'I{row}'].value == 'TOTAL':
        total_found = True
        total_row = row
        break

if total_found:
    print(f"\nTOTAL ROW (row {total_row}):")
    print(f"  I{total_row}: '{ws[f'I{total_row}'].value}' (should be 'TOTAL')")
    print(f"  K{total_row}: '{ws[f'K{total_row}'].value}' (should have SUM formula)")
else:
    print("\n❌ TOTAL not found in column I")

# Check Labor/Material rows
labor_found = False
material_found = False
for row in range(1, 30):
    if ws[f'D{row}'].value == 'LABOR:':
        labor_found = True
        print(f"\nLABOR ROW (row {row}):")
        print(f"  D{row}: '{ws[f'D{row}'].value}' (should be 'LABOR:')")
        print(f"  G{row}: '{ws[f'G{row}'].value}' (should be 'Will be 43% of total amount')")

    if ws[f'D{row}'].value == 'MATERIAL:':
        material_found = True
        print(f"\nMATERIAL ROW (row {row}):")
        print(f"  D{row}: '{ws[f'D{row}'].value}' (should be 'MATERIAL:')")
        print(f"  G{row}: '{ws[f'G{row}'].value}' (should be 'will be 28% of total amount' with lowercase 'will')")

if not labor_found:
    print("\n❌ LABOR: not found in column D")
if not material_found:
    print("\n❌ MATERIAL: not found in column D")

# Check column widths
print("\nCOLUMN WIDTHS:")
print(f"  A: {ws.column_dimensions['A'].width:.1f} (should be ~4)")
print(f"  B: {ws.column_dimensions['B'].width:.1f} (should be ~7)")
print(f"  C: {ws.column_dimensions['C'].width:.1f} (should be ~7)")
print(f"  D: {ws.column_dimensions['D'].width:.1f} (should be ~12)")
print(f"  H: {ws.column_dimensions['H'].width:.1f} (should be ~16)")
print(f"  K: {ws.column_dimensions['K'].width:.1f} (should be ~12)")

print("\n" + "=" * 60)
print("VERIFICATION COMPLETE")
print("=" * 60)

wb.close()

# Clean up
import os
os.remove(test_filename)
if os.path.exists(output_path):
    os.remove(output_path)