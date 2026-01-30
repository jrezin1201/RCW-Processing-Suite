#!/usr/bin/env python3
"""Test the metadata extraction script with various data formats."""

import openpyxl
from openpyxl import Workbook
import os

# Test cases with different formats
test_cases = [
    {
        "name": "test1_standard.xlsx",
        "B3": "OCH - Arches - 1157451",
        "B5": "PH07 - HS 44-49, 144-142",
        "expected_project": "Arches",
        "expected_phase": 7
    },
    {
        "name": "test2_double_digit.xlsx",
        "B3": "ABC - WestView Heights - 987654",
        "B5": "PH11 - HS 100-150",
        "expected_project": "WestView Heights",
        "expected_phase": 11
    },
    {
        "name": "test3_spaces.xlsx",
        "B3": "XYZ - Mountain Ridge Estates - 555555",
        "B5": "PH03 - Building A",
        "expected_project": "Mountain Ridge Estates",
        "expected_phase": 3
    },
    {
        "name": "test4_phase_twenty.xlsx",
        "B3": "DEF - Sunset Valley - 111222",
        "B5": "PH20 - Units 200-250",
        "expected_project": "Sunset Valley",
        "expected_phase": 20
    }
]

print("Creating test files and extracting metadata...\n")
print("=" * 60)

for test_case in test_cases:
    # Create test file
    wb = Workbook()
    ws = wb.active
    ws['B3'] = test_case['B3']
    ws['B5'] = test_case['B5']
    wb.save(test_case['name'])
    wb.close()

    print(f"\nTest File: {test_case['name']}")
    print(f"  B3: {test_case['B3']}")
    print(f"  B5: {test_case['B5']}")

    # Run extraction script
    from extract_project_metadata import extract_project_metadata

    project_name, phase_number = extract_project_metadata(test_case['name'])

    # Check results
    print(f"  Extracted: Project: {project_name} | Phase: {phase_number}")
    print(f"  Expected:  Project: {test_case['expected_project']} | Phase: {test_case['expected_phase']}")

    # Verify correctness
    if project_name == test_case['expected_project'] and phase_number == test_case['expected_phase']:
        print("  ✅ PASSED")
    else:
        print("  ❌ FAILED")

    # Clean up test file
    os.remove(test_case['name'])

print("\n" + "=" * 60)
print("All tests completed!")