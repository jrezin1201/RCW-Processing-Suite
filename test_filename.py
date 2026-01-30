#!/usr/bin/env python3
"""Test that output files are now named contracts_forms instead of lennar_summary."""
import openpyxl
from openpyxl import Workbook
import os
from pathlib import Path

# Create test data
wb = Workbook()
ws = wb.active

ws['B3'] = "OCH - TestProject - 123456"
ws['B5'] = "PH07 - HS 0100-0105"

test_data = [
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["Lot/Block", "Plan", "Elevation", "Task", "Task Start Date", "Subtotal"],
    ["44", "2", "B", "Prime Exterior", "2024-01-01", "100.00"],
    ["45", "3", "A", "Paint Exterior", "2024-01-02", "150.00"],
]

for row_idx, row_data in enumerate(test_data[6:], 7):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

test_filename = "test_filename_input.xlsx"
wb.save(test_filename)
wb.close()

print(f"Created test input file: {test_filename}")

# Process the file
from app.services.parser_lennar import parse_lennar_export
from app.services.aggregator import aggregate_data
from app.services.excel_writer import write_summary_excel

parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(test_filename)
summary_rows, qa_report = aggregate_data(parsed_rows, qa_meta)

# Use a specific job_id for testing
test_job_id = "test_filename_check"
output_path = write_summary_excel(
    summary_rows, qa_report, test_job_id,
    phase=phase, project_name=project_name, house_string=house_string
)

print(f"\n=== Filename Check ===")
print(f"Output path: {output_path}")

# Extract just the filename
filename = Path(output_path).name
print(f"Output filename: {filename}")

# Check if it starts with contracts_forms
if filename.startswith("contracts_forms_"):
    print("✅ SUCCESS: File now starts with 'contracts_forms_'")
else:
    print(f"❌ FAILED: File does not start with 'contracts_forms_'. Starts with: {filename[:20]}")

# Clean up test files
os.remove(test_filename)
if os.path.exists(output_path):
    os.remove(output_path)
    print("\nTest files cleaned up.")