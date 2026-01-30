#!/usr/bin/env python3
"""Create a test Excel file with project metadata for demonstration."""

import openpyxl
from openpyxl import Workbook

# Create a new workbook
wb = Workbook()
ws = wb.active

# Set cell B3 with project name format
ws['B3'] = "OCH - Arches - 1157451"

# Set cell B5 with phase format
ws['B5'] = "PH07 - HS 44-49, 144-142"

# Save the test file
test_file = "test_metadata.xlsx"
wb.save(test_file)
wb.close()

print(f"✅ Created test file: {test_file}")
print(f"   B3 contains: OCH - Arches - 1157451")
print(f"   B5 contains: PH07 - HS 44-49, 144-142")