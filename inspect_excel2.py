"""Inspect an Excel file using openpyxl."""
import openpyxl

filepath = 'data/uploads/ee2ccca9-177a-49ff-8da7-1c90187139cd.xlsx'

# Open the workbook
wb = openpyxl.load_workbook(filepath, data_only=True)
ws = wb.active

print(f"File: {filepath}")
print(f"Active sheet: {ws.title}")
print(f"Dimensions: {ws.max_row} rows, {ws.max_column} columns")

print("\n=== First 10 Rows ===")
for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True), 1):
    # Skip empty rows
    if row and any(cell is not None for cell in row):
        print(f"\nRow {row_idx}:")
        for col_idx, cell in enumerate(row[:8], 1):  # Show first 8 columns
            if cell is not None:
                print(f"  Col {col_idx}: '{cell}'")

print("\n=== Checking for Lennar Headers ===")
required = ["lot/block", "plan", "task", "task start date"]
found_headers = []

# Check first 20 rows for headers
for row_idx in range(1, min(21, ws.max_row + 1)):
    row = ws[row_idx]
    row_values = [str(cell.value).lower() if cell.value else "" for cell in row]
    row_text = ' '.join(row_values)

    matches = []
    for header in required:
        if header in row_text:
            matches.append(header)

    if matches:
        print(f"Row {row_idx}: Found potential headers: {matches}")
        if len(matches) >= 2:
            print(f"  Full row content (first 8 cols):")
            for col_idx, cell in enumerate(row[:8], 1):
                if cell.value:
                    print(f"    Col {col_idx}: '{cell.value}'")

wb.close()

print("\n=== Required Headers for Lennar Format ===")
print("The service needs these exact column headers:")
print("  1. 'Lot/Block' - for lot identification")
print("  2. 'Plan' - for plan type")
print("  3. 'Task' - for task description")
print("  4. 'Task Start Date' - for scheduling")
print("\nYour file appears to be missing some or all of these headers.")