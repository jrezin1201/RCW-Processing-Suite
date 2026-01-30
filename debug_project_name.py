#!/usr/bin/env python3
"""Debug script to investigate why project name is not being extracted."""
import sys
import openpyxl
import pandas as pd

def debug_extraction(file_path):
    """Debug the extraction of project metadata."""
    print("=" * 60)
    print("DEBUGGING PROJECT NAME EXTRACTION")
    print("=" * 60)
    print(f"\nFile: {file_path}")

    # Try openpyxl first, then pandas
    ws = None
    df = None

    try:
        # Try loading with openpyxl (for .xlsx)
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        print("Format: XLSX (using openpyxl)")
    except:
        # Try with pandas (for .xls)
        try:
            df = pd.read_excel(file_path)
            print("Format: XLS (using pandas)")
        except Exception as e:
            print(f"❌ Could not load file: {e}")
            return

    try:
        print("\n1. RAW CELL VALUES:")

        # Check cell B3
        if ws:  # openpyxl format
            b3_value = ws['B3'].value
        else:  # pandas format
            # B3 is row 2, column 1 in 0-indexed pandas
            b3_value = df.iloc[2, 1] if len(df) > 2 and len(df.columns) > 1 else None

        print(f"   B3 raw value: '{b3_value}'")
        print(f"   B3 type: {type(b3_value)}")

        # Check if it's None or empty
        if b3_value is None:
            print("   ⚠️ B3 is None/empty!")
            # Check surrounding cells
            print("\n   Checking surrounding cells:")
            if ws:
                for row in range(1, 6):
                    for col in ['A', 'B', 'C', 'D']:
                        cell_value = ws[f'{col}{row}'].value
                        if cell_value:
                            print(f"   {col}{row}: '{str(cell_value)[:50]}'")
            else:
                for row in range(min(5, len(df))):
                    for col in range(min(4, len(df.columns))):
                        cell_value = df.iloc[row, col]
                        if pd.notna(cell_value):
                            print(f"   Row {row+1}, Col {chr(65+col)}: '{str(cell_value)[:50]}'")

        # Try to extract project name
        print("\n2. EXTRACTION ATTEMPT:")
        if b3_value:
            b3_str = str(b3_value)
            print(f"   B3 as string: '{b3_str}'")

            # Try splitting by hyphen
            parts = b3_str.split('-')
            print(f"   Split by '-': {parts}")

            if len(parts) >= 2:
                project_name = parts[1].strip()
                print(f"   ✅ Extracted project name: '{project_name}'")
            else:
                print(f"   ❌ Could not extract - not enough parts")
        else:
            print("   ❌ Cannot extract from empty/None value")

        # Check B5 for phase
        print("\n3. PHASE EXTRACTION (B5):")
        if ws:
            b5_value = ws['B5'].value
        else:
            # B5 is row 4, column 1 in 0-indexed pandas
            b5_value = df.iloc[4, 1] if len(df) > 4 and len(df.columns) > 1 else None
        print(f"   B5 raw value: '{b5_value}'")
        print(f"   B5 type: {type(b5_value)}")

        # Also check if data might be in different rows
        print("\n4. SCANNING FOR PROJECT DATA PATTERN:")
        found_project = False

        if ws:
            for row in range(1, min(20, ws.max_row + 1)):
                for col in range(1, 5):  # Check columns A-D
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and ' - ' in str(cell_value):
                        parts = str(cell_value).split('-')
                        if len(parts) >= 3:  # Looking for pattern like "XXX - ProjectName - Number"
                            print(f"   Found potential project data at row {row}, col {col}: '{cell_value}'")
                            found_project = True
        else:
            for row in range(min(20, len(df))):
                for col in range(min(4, len(df.columns))):
                    cell_value = df.iloc[row, col]
                    if pd.notna(cell_value) and ' - ' in str(cell_value):
                        parts = str(cell_value).split('-')
                        if len(parts) >= 3:
                            print(f"   Found potential project data at row {row+1}, col {chr(65+col)}: '{cell_value}'")
                            found_project = True

        if not found_project:
            print("   No project pattern found in first 20 rows")

        # Test with actual extraction functions
        print("\n5. USING ACTUAL PARSER FUNCTIONS:")
        from app.services.parser_lennar import parse_lennar_export

        # Use the full parser to see what it extracts
        parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(file_path)

        print(f"   Full parser extracted:")
        print(f"     - project_name: '{project_name}'")
        print(f"     - phase: '{phase}'")
        print(f"     - house_string: '{house_string}'")

        if ws:
            wb.close()

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    # Check for existing uploaded files to test
    from pathlib import Path

    uploads_dir = Path("data/uploads")
    if uploads_dir.exists():
        files = list(uploads_dir.glob("*.xlsx")) + list(uploads_dir.glob("*.xls"))

        if files:
            print("\nFound uploaded files:")
            for i, f in enumerate(files[-5:], 1):  # Show last 5 files
                print(f"{i}. {f.name}")

            if len(sys.argv) > 1:
                file_path = sys.argv[1]
            else:
                # Test with the most recent file
                file_path = str(files[-1])
                print(f"\nTesting with most recent: {file_path}")

            debug_extraction(file_path)
        else:
            print("No uploaded files found. Upload a file through the web interface first.")
    else:
        print("Upload directory doesn't exist.")