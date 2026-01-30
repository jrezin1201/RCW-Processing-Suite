#!/usr/bin/env python3
"""Test that the final output now correctly shows the project name."""
from pathlib import Path

# Test with the most recent uploaded file
uploaded_files = list(Path("data/uploads").glob("*.xlsx"))
if uploaded_files:
    test_file = str(uploaded_files[-1])
    print(f"Testing with: {test_file}")

    from app.services.parser_lennar import parse_lennar_export
    from app.services.aggregator import aggregate_data
    from app.services.excel_writer import write_summary_excel
    import openpyxl

    # Parse the file
    parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(test_file)

    print("\n=== EXTRACTED METADATA ===")
    print(f"Project Name: {project_name}")
    print(f"Phase: {phase}")
    print(f"House String: {house_string}")

    # Generate output
    summary_rows, qa_report = aggregate_data(parsed_rows, qa_meta)
    output_path = write_summary_excel(
        summary_rows, qa_report, "test_final",
        phase=phase, project_name=project_name, house_string=house_string
    )

    print(f"\n=== OUTPUT FILE CREATED ===")
    print(f"Path: {output_path}")

    # Check the output
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    print("\n=== OUTPUT VERIFICATION ===")
    print(f"B1 (Project Name): '{ws['B1'].value}'")
    print(f"G1 (Phase): '{ws['G1'].value}'")
    print(f"H1 (House String): '{ws['H1'].value}'")

    if "Arches" in str(ws['B1'].value):
        print("\n✅ SUCCESS! Project name 'Arches' is now displayed in the output!")
    else:
        print(f"\n❌ Project name still missing. B1 = '{ws['B1'].value}'")

    wb.close()
else:
    print("No uploaded files found.")