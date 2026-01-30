#!/usr/bin/env python3
"""
Standalone script to extract project metadata from Excel spreadsheets.

This script extracts:
1. Project Name from cell B3 (text between first and second hyphen)
2. Phase Number from cell B5 (digits after "PH")
3. House String from cell B5 (text after "PH## - ")
"""

import openpyxl
import re
import sys


def extract_metadata_from_excel(filepath):
    """
    Extract project metadata from an Excel file.

    Args:
        filepath: Path to the Excel file

    Returns:
        Dictionary with project_name, phase_number, and house_string
    """
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        metadata = {}

        # 1. Extract Project Name from B3
        cell_b3 = ws.cell(row=3, column=2).value  # B3
        if cell_b3:
            # Extract text between first and second hyphen
            parts = str(cell_b3).split('-')
            if len(parts) >= 2:
                metadata['project_name'] = parts[1].strip()
            else:
                metadata['project_name'] = None
        else:
            metadata['project_name'] = None

        # 2. Extract Phase Number and House String from B5
        cell_b5 = ws.cell(row=5, column=2).value  # B5
        if cell_b5:
            # Extract phase number (PHxx pattern)
            phase_match = re.search(r'PH(\d{2})', str(cell_b5), re.IGNORECASE)
            if phase_match:
                # Convert to integer to remove leading zeros
                phase_num = int(phase_match.group(1))
                metadata['phase_number'] = phase_num
            else:
                metadata['phase_number'] = None

            # Extract house string (text after "PH## - ")
            house_match = re.search(r'PH\d{2}\s*-\s*(.+)', str(cell_b5), re.IGNORECASE)
            if house_match:
                metadata['house_string'] = house_match.group(1).strip()
            else:
                metadata['house_string'] = None
        else:
            metadata['phase_number'] = None
            metadata['house_string'] = None

        wb.close()
        return metadata

    except Exception as e:
        print(f"Error reading file: {e}")
        return None


def main():
    """Main function to demonstrate metadata extraction."""

    # Check if filepath was provided
    if len(sys.argv) < 2:
        print("Usage: python extract_metadata.py <excel_file>")
        print("\nCreating test file for demonstration...")

        # Create a test Excel file
        wb = openpyxl.Workbook()
        ws = wb.active

        # Add test data
        ws['B3'] = "OCH - Arches - 1157451"  # Project name example
        ws['B5'] = "PH07 - HS 0044-0049, 0144-0142"  # Phase and house string example

        # Add some dummy data to make it look like a real sheet
        ws['A7'] = "Lot/Block"
        ws['B7'] = "Plan"
        ws['C7'] = "Task"

        test_filename = "test_metadata_example.xlsx"
        wb.save(test_filename)
        wb.close()

        print(f"Created test file: {test_filename}")
        filepath = test_filename
    else:
        filepath = sys.argv[1]

    print(f"\nExtracting metadata from: {filepath}")
    print("=" * 50)

    # Extract metadata
    metadata = extract_metadata_from_excel(filepath)

    if metadata:
        # Format output as requested
        project_name = metadata['project_name'] or "Not found"
        phase_number = metadata['phase_number'] if metadata['phase_number'] is not None else "Not found"
        house_string = metadata['house_string'] or "Not found"

        print(f"\nProject: {project_name} | Phase: {phase_number}")
        print(f"\nDetailed Results:")
        print(f"  Project Name: {project_name}")
        print(f"  Phase Number: {phase_number}")
        print(f"  House String: {house_string}")

        print("\n" + "=" * 50)
        print("Extraction Logic Explanation:")
        print("1. Project Name (B3): Extracted word between first and second hyphen")
        print("2. Phase Number (B5): Found 'PH' pattern and extracted digits, removed leading zeros")
        print("3. House String (B5): Extracted text after 'PH## - ' pattern")
    else:
        print("Failed to extract metadata from file.")


if __name__ == "__main__":
    main()