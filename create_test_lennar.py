"""Create a test Lennar Excel file with proper format."""
import openpyxl
from datetime import datetime

# Create new workbook
wb = openpyxl.Workbook()
ws = wb.active

# Add headers (Lennar format)
headers = [
    "Lot/Block", "Plan", "Elevation", "Swing", "Task Start Date",
    "Task", "Job Subtotal", "Tax", "Total"
]

# Add headers in row 3 (leaving some blank rows like real Lennar files)
for col, header in enumerate(headers, 1):
    ws.cell(row=3, column=col, value=header)

# Add sample data rows
test_data = [
    # Lot 101
    ["101", "Plan A", "Elev 1", None, datetime(2024, 1, 15), "Painting - Exterior Prime", 1500.00, 120.00, 1620.00],
    ["101", "Plan A", "Elev 1", None, datetime(2024, 1, 16), "Painting - Interior Walls", 2000.00, 160.00, 2160.00],
    ["101", "Plan A", "Elev 1", None, datetime(2024, 1, 17), "Painting - Exterior [UA]", 1800.00, 144.00, 1944.00],

    # Lot 102
    ["102", "Plan B", "Elev 2", None, datetime(2024, 1, 15), "Painting - Exterior", 1200.00, 96.00, 1296.00],
    ["102", "Plan B", "Elev 2", None, datetime(2024, 1, 16), "Painting - Interior Ceilings", 1500.00, 120.00, 1620.00],
    ["102", "Plan B", "Elev 2", None, datetime(2024, 1, 17), "Painting - Ext Prime Coat", 1600.00, 128.00, 1728.00],

    # Lot 103
    ["103", "Plan A", "Elev 3", None, datetime(2024, 1, 18), "Exterior Stucco Prime", 2200.00, 176.00, 2376.00],
    ["103", "Plan A", "Elev 3", None, datetime(2024, 1, 19), "Interior Paint - Bedrooms", 1800.00, 144.00, 1944.00],
    ["103", "Plan A", "Elev 3", None, datetime(2024, 1, 20), "Exterior UA Application", 2100.00, 168.00, 2268.00],

    # Lot 104 - with some unmapped tasks
    ["104", "Plan C", "Elev 1", None, datetime(2024, 1, 21), "Cleaning Services", 500.00, 40.00, 540.00],
    ["104", "Plan C", "Elev 1", None, datetime(2024, 1, 22), "Painting - Exterior Finish", 1400.00, 112.00, 1512.00],
    ["104", "Plan C", "Elev 1", None, datetime(2024, 1, 23), "Touch Up Work", 300.00, 24.00, 324.00],
]

# Write data starting from row 4
for row_idx, row_data in enumerate(test_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        # Format money columns
        if col_idx in [7, 8, 9]:  # Subtotal, Tax, Total columns
            cell.number_format = '"$"#,##0.00'

# Save the file
wb.save("lennar_test_data.xlsx")
wb.close()

print("Created lennar_test_data.xlsx with proper Lennar format")
print(f"- {len(test_data)} data rows")
print("- Headers: " + ", ".join(headers))
print("\nTasks include:")
for row in test_data:
    print(f"  - {row[5]}")