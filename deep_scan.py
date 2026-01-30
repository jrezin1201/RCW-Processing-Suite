#!/usr/bin/env python3
"""
Deep scan to find where actual data starts and identify column patterns.
"""

import sys
import openpyxl
import re

def deep_scan(filepath):
    """Scan entire Excel file to find data patterns."""

    print("="*60)
    print(f"DEEP SCANNING FILE: {filepath}")
    print("="*60)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        print(f"\n📊 File Info:")
        print(f"  Sheet: {ws.title}")
        print(f"  Total Rows: {ws.max_row}")
        print(f"  Total Columns: {ws.max_column}")

        # Find first non-empty row with substantial data
        print("\n🔍 FINDING DATA START:")
        print("-"*60)

        data_start_row = None
        job_number_pattern = re.compile(r'^\s*(\d{4})\b')

        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num > 100:  # Check first 100 rows
                break

            # Count non-empty cells
            non_empty = sum(1 for cell in row if cell is not None and str(cell).strip())

            if non_empty >= 3:  # Row has at least 3 non-empty cells
                print(f"\nRow {row_num} has {non_empty} non-empty cells:")
                for col_idx, value in enumerate(row[:13] if row else []):
                    if value is not None and str(value).strip():
                        col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                        value_str = str(value)[:50]  # Truncate long values
                        print(f"  {col_letter} (idx {col_idx}): {repr(value_str)}")

                        # Check if this looks like a job number
                        if job_number_pattern.match(str(value)):
                            print(f"    → Possible job number: {job_number_pattern.match(str(value)).group(1)}")

        # Search for "Location Total" or similar patterns anywhere
        print("\n🔍 SEARCHING FOR TOTAL/SUBTOTAL PATTERNS:")
        print("-"*60)

        total_patterns = ['total', 'subtotal', 'sum', 'hours']
        found_totals = False

        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num > 200:  # Check first 200 rows
                break

            row_text = " ".join([str(x).lower() for x in row if x is not None])
            for pattern in total_patterns:
                if pattern in row_text:
                    print(f"\nRow {row_num} contains '{pattern}':")
                    for col_idx, value in enumerate(row[:13] if row else []):
                        if value is not None:
                            col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                            print(f"  {col_letter}: {str(value)[:50]}")
                    found_totals = True
                    break

            if found_totals and row_num > 50:  # Stop after finding some examples
                break

        # Look for numeric patterns that might be hours
        print("\n🔍 SEARCHING FOR NUMERIC PATTERNS (possible hours):")
        print("-"*60)

        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num < 10 or row_num > 50:  # Skip early rows, check rows 10-50
                continue

            numeric_cells = []
            for col_idx, value in enumerate(row if row else []):
                if value is not None and isinstance(value, (int, float)):
                    if 0 < float(value) < 100:  # Reasonable hours range
                        col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                        numeric_cells.append(f"{col_letter}={value}")

            if numeric_cells:
                print(f"Row {row_num}: {', '.join(numeric_cells)}")

        # Find columns with 4-digit numbers
        print("\n🔍 COLUMNS WITH 4-DIGIT NUMBERS (possible job numbers):")
        print("-"*60)

        job_cols = {}
        for row_num, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_num > 100:
                break

            for col_idx, value in enumerate(row if row else []):
                if value is not None:
                    match = job_number_pattern.match(str(value))
                    if match:
                        col_letter = chr(65 + col_idx) if col_idx < 26 else f"Col{col_idx}"
                        if col_letter not in job_cols:
                            job_cols[col_letter] = []
                        job_cols[col_letter].append((row_num, match.group(1)))

        for col, jobs in job_cols.items():
            print(f"\nColumn {col}: Found {len(jobs)} job numbers")
            for row, job in jobs[:5]:  # Show first 5
                print(f"  Row {row}: {job}")

    except Exception as e:
        print(f"\n❌ ERROR: {e}")

if __name__ == "__main__":
    if len(sys.argv) > 1:
        deep_scan(sys.argv[1])
    else:
        print("Usage: python deep_scan.py your_file.xlsx")