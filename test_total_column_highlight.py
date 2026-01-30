#!/usr/bin/env python3
"""Test that Total column (K) has yellow highlighting for all cells."""
import openpyxl
from openpyxl import Workbook

# Create test data similar to the image
wb = Workbook()
ws = wb.active

# Add metadata
ws['B3'] = "OCH - Arches - 123456"  # Project name from image
ws['B5'] = "PH07 - HS 44-49, 144-142"  # Phase and house string from image

# Add test data matching the image
test_data = [
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["Lot/Block", "Plan", "Elevation", "Task", "Task Start Date", "Subtotal"],
    ["44", "2B", "B", "Prime Exterior", "2024-01-01", "1543.50"],
    ["45", "3A", "A", "Exterior", "2024-01-02", "1516.49"],
    ["46", "1B", "C", "Exterior UA", "2024-01-03", "1056.00"],
    ["47", "2C", "D", "Interior", "2024-01-04", "4115.99"],
    ["48", "3B", "E", "", "", ""],
    ["49", "1C", "F", "", "", ""],
    ["142", "3A", "G", "", "", ""],
    ["143", "2B", "H", "", "", ""],
    ["144", "3C", "I", "", "", ""],
]

for row_idx, row_data in enumerate(test_data[6:], 7):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

test_filename = "test_total_highlight_input.xlsx"
wb.save(test_filename)
wb.close()

print(f"Created test input file: {test_filename}")

# Process the file
from app.services.parser_lennar import parse_lennar_export
from app.services.aggregator import aggregate_data
from app.services.excel_writer import write_summary_excel

parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(test_filename)
summary_rows, qa_report = aggregate_data(parsed_rows, qa_meta)
output_path = write_summary_excel(
    summary_rows, qa_report, "test_total_highlight",
    phase=phase, project_name=project_name, house_string=house_string
)

print(f"Output written to: {output_path}")

# Check the yellow highlighting in column K
wb = openpyxl.load_workbook(output_path)
ws = wb.active

print("\n=== Total Column (K) Highlighting Analysis ===")

# Check cells from row 3 (header) to row 13 (total row)
yellow_count = 0
non_yellow_count = 0

for row in range(3, 14):  # Rows 3 to 13
    cell = ws[f"K{row}"]

    has_yellow = False
    if cell.fill and cell.fill.start_color:
        fill_color = cell.fill.start_color.index
        if fill_color == "FFFF00" or fill_color == "00FFFF00":
            has_yellow = True
            yellow_count += 1
        else:
            non_yellow_count += 1
    else:
        non_yellow_count += 1

    value_str = f"'{cell.value}'" if cell.value else "empty"
    status = "✅ Yellow" if has_yellow else "❌ No yellow"
    print(f"  K{row}: {status} (value: {value_str})")

print(f"\n=== Summary ===")
print(f"Yellow cells: {yellow_count}")
print(f"Non-yellow cells: {non_yellow_count}")

if yellow_count >= 10:  # Should have yellow from row 3 to at least row 13
    print("✅ Total column has consistent yellow highlighting as shown in the image!")
else:
    print("⚠️ Total column may need more yellow highlighting")

# Also check specific formatting requirements
print("\n=== Other Formatting Checks ===")

# Check merged cells B1:C1
if "B1:C1" in str(ws.merged_cells.ranges):
    print("✅ B1:C1 are merged for Project Name")
else:
    print("❌ B1:C1 are not merged")

# Check TOTAL label position (should be in J)
j_total = ws["J13"].value if ws["J13"].value else ws["J12"].value if ws["J12"].value else ws["J11"].value
if j_total and "TOTAL" in str(j_total):
    print("✅ TOTAL label is in column J")
else:
    print("❌ TOTAL label is not in column J")

# Check Project Name in B1
if ws["B1"].value and "Arches" in str(ws["B1"].value):
    print(f"✅ Project Name displayed: {ws['B1'].value}")
else:
    print("❌ Project Name not properly displayed")

# Check Phase in G1
if ws["G1"].value and "Phase" in str(ws["G1"].value):
    print(f"✅ Phase displayed: {ws['G1'].value}")
else:
    print("❌ Phase not properly displayed")

wb.close()