"""Generate warning notices for timekeeping violations from an Exception List export.

Exposes the pure parsing and workbook-building logic used by the Missed Clock-In
FastAPI routes.
"""

from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.page import PageMargins

SKIP_NAMES = {
    "Day", "RC Wendt Painting", "Employee Missed Ins/Outs",
    "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday",
}

TARGET_ERRORS = {"Missing Start Touch", "Missing Stop Touch"}
TRACKED_ERRORS = TARGET_ERRORS | {"Clocked In Twice"}

COLUMN_WIDTHS = {
    "A": 12.43, "B": 15.71, "C": 13.0, "D": 13.0,
    "E": 12.57, "F": 13.0, "G": 39.57, "H": 7.0,
}

_INVALID_SHEET_CHARS = set(r':\/?*[]')


def is_employee_name_row(row):
    c0 = row[0]
    c2 = row[2] if len(row) > 2 else None
    c9 = row[9] if len(row) > 9 else None
    if not isinstance(c0, str) or not c0.strip():
        return False
    if pd.notna(c2) or pd.notna(c9):
        return False
    text = c0.strip()
    if text in SKIP_NAMES:
        return False
    if text.startswith("Time Records"):
        return False
    return True


def parse_exception_list(path: Path):
    df = pd.read_excel(path, header=None)
    records = []
    current_employee = None
    for _, row in df.iterrows():
        row = row.tolist()
        while len(row) < 10:
            row.append(None)

        if is_employee_name_row(row):
            current_employee = row[0].strip()
            continue

        err = row[9]
        if isinstance(err, str) and err.strip() in TRACKED_ERRORS and current_employee:
            records.append({
                "employee": current_employee,
                "date": row[2],
                "start_time": row[3],
                "start_location": row[5],
                "stop_time": row[6],
                "stop_location": row[7],
                "error": err.strip(),
            })
    return records


def format_date(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    if isinstance(val, datetime):
        return val.strftime("%m/%d/%Y")
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%m/%d/%Y").strftime("%m/%d/%Y")
        except ValueError:
            return val.strip()
    return str(val)


def parse_date(val):
    """Return a datetime object for the date value, or None."""
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return None
    if isinstance(val, datetime):
        return val.replace(hour=0, minute=0, second=0, microsecond=0)
    if isinstance(val, str):
        try:
            return datetime.strptime(val.strip(), "%m/%d/%Y")
        except ValueError:
            return None
    return None


def format_time(val):
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    return str(val).strip()


def description_lines(error):
    if error == "Missing Start Touch":
        return (
            "Employee did not clock in but was paid 8 regular hours after calling the employee and foreman",
            "to verify.",
            "El empleado no fichó, pero se le pagaron 8 horas regulares después de llamar al empleado y",
            "al foreman para verificar.",
        )
    return (
        "Employee did not clock out but was paid 8 regular hours after calling the employee and foreman",
        "to verify.",
        "El empleado no fichó la salida, pero se le pagaron 8 horas regulares después de llamar al empleado y",
        "al foreman para verificar.",
    )


def apply_notice_borders(ws, R):
    """Apply all borders for a single notice starting at row R (1-indexed)."""
    thin = Side(style='thin')

    def set_border(row, col, top=None, bottom=None, left=None, right=None):
        c = ws.cell(row=row, column=col)
        c.border = Border(
            top=top or Side(),
            bottom=bottom or Side(),
            left=left or Side(),
            right=right or Side(),
        )

    # Section 1: Employee name underline (row R+4, cols C-F)
    for c in (3, 4, 5, 6):
        set_border(R + 4, c, bottom=thin)

    # Section 2: Policy bar (row R+7, cols A-H)
    for c in range(1, 9):
        set_border(R + 7, c, bottom=thin)

    # Section 3: Date & Description box (rows R+8 to R+12)
    # Row R+8
    set_border(R + 8, 1, top=thin, left=thin, right=thin)
    set_border(R + 8, 2, top=thin, right=thin)
    for c in (3, 4, 5, 6):
        set_border(R + 8, c, top=thin, bottom=thin)
    set_border(R + 8, 7, top=thin)
    set_border(R + 8, 8, top=thin, right=thin)

    # Row R+9
    set_border(R + 9, 1, left=thin, right=thin)
    set_border(R + 9, 2, right=thin)
    set_border(R + 9, 3, bottom=thin, left=thin)
    for c in (4, 5, 6, 7):
        set_border(R + 9, c, bottom=thin)
    set_border(R + 9, 8, bottom=thin, right=thin)

    # Row R+10
    set_border(R + 10, 1, left=thin)
    set_border(R + 10, 2, right=thin)
    set_border(R + 10, 3, bottom=thin, left=thin)
    for c in (4, 5, 6, 7):
        set_border(R + 10, c, bottom=thin)
    set_border(R + 10, 8, bottom=thin, right=thin)

    # Row R+11
    set_border(R + 11, 1, left=thin)
    set_border(R + 11, 2, right=thin)
    set_border(R + 11, 3, top=thin, left=thin, right=thin)
    for c in (4, 5, 6, 7):
        set_border(R + 11, c, top=thin)
    set_border(R + 11, 8, top=thin, right=thin)

    # Row R+12
    set_border(R + 12, 1, bottom=thin, left=thin)
    set_border(R + 12, 2, bottom=thin, right=thin)
    set_border(R + 12, 3, bottom=thin, left=thin, right=thin)
    for c in (4, 5, 6, 7):
        set_border(R + 12, c, bottom=thin)
    set_border(R + 12, 8, bottom=thin, right=thin)

    # Section 4: Fill-out box (rows R+16 to R+23)
    # Row R+16 (Job #)
    set_border(R + 16, 2, top=thin, left=thin)
    set_border(R + 16, 3, top=thin)
    set_border(R + 16, 4, top=thin, bottom=thin)
    set_border(R + 16, 5, top=thin, bottom=thin)
    set_border(R + 16, 6, top=thin)
    set_border(R + 16, 7, top=thin)
    set_border(R + 16, 8, top=thin, right=thin)

    # Row R+17 (Clock In)
    set_border(R + 17, 2, left=thin)
    set_border(R + 17, 4, bottom=thin)
    set_border(R + 17, 8, right=thin)

    # Row R+18 (Clock Out)
    set_border(R + 18, 2, left=thin)
    set_border(R + 18, 4, top=thin, bottom=thin)
    set_border(R + 18, 8, right=thin)

    # Row R+19 (Hours Worked)
    set_border(R + 19, 2, left=thin)
    set_border(R + 19, 4, top=thin, bottom=thin)
    set_border(R + 19, 8, right=thin)

    # Row R+20 (Explanation)
    set_border(R + 20, 2, left=thin)
    set_border(R + 20, 4, top=thin, bottom=thin)
    for c in (5, 6, 7):
        set_border(R + 20, c, bottom=thin)
    set_border(R + 20, 8, right=thin)

    # Rows R+21 and R+22 (Explanation continuation)
    for r_off in (21, 22):
        set_border(R + r_off, 2, left=thin)
        for c in (4, 5, 6, 7):
            set_border(R + r_off, c, top=thin, bottom=thin)
        set_border(R + r_off, 8, right=thin)

    # Row R+23 (Box bottom)
    set_border(R + 23, 2, bottom=thin, left=thin)
    set_border(R + 23, 3, bottom=thin)
    for c in (4, 5, 6, 7):
        set_border(R + 23, c, top=thin, bottom=thin)
    set_border(R + 23, 8, bottom=thin, right=thin)

    # Section 5: Signature lines
    # Row R+26 (Employee Signature)
    for c in (3, 4, 5):
        set_border(R + 26, c, bottom=thin)
    set_border(R + 26, 7, bottom=thin)

    # Row R+28 (Foreman Signature)
    for c in (3, 4, 5):
        set_border(R + 28, c, bottom=thin)
    set_border(R + 28, 7, bottom=thin)

    # Row R+30 (Superintendent Signature)
    for c in (4, 5, 6, 7):
        set_border(R + 30, c, bottom=thin)


def write_notice(ws, start_row, rec):
    """Write one 34-row warning notice block starting at start_row (1-indexed)."""
    def cell(r, col):
        return ws.cell(row=start_row + r - 1, column=col)

    def merge(r1, c1, r2, c2):
        ws.merge_cells(
            start_row=start_row + r1 - 1, start_column=c1,
            end_row=start_row + r2 - 1, end_column=c2,
        )

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Row 1-2: Title English (B:G merged across 2 rows)
    merge(1, 2, 2, 7)
    c = cell(1, 2)
    c.value = "Warning Notice for Time Keeping Policy"
    c.font = Font(name="Calibri", size=28, bold=True)
    c.alignment = center_wrap

    # Row 3: Title Spanish
    merge(3, 2, 3, 7)
    c = cell(3, 2)
    c.value = " Aviso de Advertencia por Política de Control de Tiempo"
    c.font = Font(name="Calibri", size=14, bold=True)
    c.alignment = center

    # Row 5: Employee label + name
    merge(5, 1, 5, 2)
    c = cell(5, 1)
    c.value = "Employee / Empleado:"
    c.font = Font(name="Calibri", size=11, bold=True)
    c.alignment = center

    merge(5, 3, 5, 6)
    c = cell(5, 3)
    c.value = rec["employee"]
    c.font = Font(name="Calibri", size=11)
    c.alignment = center

    # Row 8: Policy heading
    merge(8, 1, 8, 8)
    c = cell(8, 1)
    c.value = ("Not Following Company Policy for Time Keeping / "
               "No seguir la política de la empresa para el control de tiempo")
    c.font = Font(name="Calibri", size=13.5, bold=True)
    c.alignment = center

    # Row 9: Date of incident
    merge(9, 1, 9, 2)
    c = cell(9, 1)
    c.value = "Date of Incident / Fecha del incidente:"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.alignment = center

    merge(9, 3, 9, 6)
    c = cell(9, 3)
    c.value = parse_date(rec["date"])
    c.font = Font(name="Calibri", size=11)
    c.number_format = 'M/D/YYYY'
    c.alignment = center

    # Row 10: Description label + English line 1 (NO merge on C10)
    merge(10, 1, 10, 2)
    c = cell(10, 1)
    c.value = "Description / Descripción:"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.alignment = center

    eng1, eng2, spa1, spa2 = description_lines(rec["error"])

    # C10 is intentionally not merged so text can overflow across C–H.
    c = cell(10, 3)
    c.value = eng1
    c.font = Font(name="Calibri", size=11)
    c.alignment = left

    # Row 11: English line 2
    merge(11, 3, 11, 7)
    c = cell(11, 3)
    c.value = eng2
    c.font = Font(name="Calibri", size=11)
    c.alignment = left

    # Row 12: Spanish line 1
    merge(12, 3, 12, 8)
    c = cell(12, 3)
    c.value = spa1
    c.font = Font(name="Calibri", size=10)
    c.alignment = left

    # Row 13: Spanish line 2
    merge(13, 3, 13, 8)
    c = cell(13, 3)
    c.value = spa2
    c.font = Font(name="Calibri", size=11)
    c.alignment = left

    # Row 16: Fill out heading
    c = cell(16, 2)
    c.value = "Fill Out the Following: / Completar lo siguiente:"
    c.font = Font(name="Calibri", size=11, bold=True)
    c.alignment = left

    # Row 17: Job # label; D17 intentionally left blank for manual fill-in.
    c = cell(17, 2)
    c.value = "Job # / Número de trabajo:"
    c.font = Font(name="Calibri", size=11, bold=True)
    c.alignment = left

    c = cell(17, 4)
    c.font = Font(name="Calibri", size=11)
    c.alignment = center

    # Row 18: Clock In
    merge(18, 2, 18, 3)
    c = cell(18, 2)
    c.value = "Clock In Time / Hora de entrada:"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.alignment = center

    c = cell(18, 4)
    if rec["error"] == "Missing Stop Touch":
        c.value = format_time(rec["start_time"])
    c.font = Font(name="Calibri", size=11)
    c.alignment = center

    # Row 19: Clock Out
    merge(19, 2, 19, 3)
    c = cell(19, 2)
    c.value = "Clock Out Time / Hora de salida:"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.alignment = center

    c = cell(19, 4)
    if rec["error"] == "Missing Start Touch":
        c.value = format_time(rec["stop_time"])
    c.font = Font(name="Calibri", size=11)
    c.alignment = center

    # Row 20: Hours Worked
    merge(20, 2, 20, 3)
    c = cell(20, 2)
    c.value = "Hours Worked / Horas trabajadas:"
    c.font = Font(name="Calibri", size=10, bold=True)
    c.alignment = center

    # Row 21: Explanation
    merge(21, 2, 21, 3)
    c = cell(21, 2)
    c.value = "Explanation / Explicación:"
    c.font = Font(name="Calibri", size=11, bold=True)
    c.alignment = center

    # Row 27: Employee signature
    merge(27, 1, 27, 2)
    c = cell(27, 1)
    c.value = "Employee Signature / Firma del empleado:"
    c.font = Font(name="Calibri", size=8)
    c.alignment = center

    c = cell(27, 6)
    c.value = "Date/ Fecha:"
    c.font = Font(name="Calibri", size=11)

    # Row 29: Foreman signature
    merge(29, 1, 29, 2)
    c = cell(29, 1)
    c.value = "Foreman Signature/ firma de foreman:"
    c.font = Font(name="Calibri", size=10)
    c.alignment = center

    c = cell(29, 6)
    c.value = "Date/ Fecha:"
    c.font = Font(name="Calibri", size=11)

    # Row 31: Superintendent signature (A31:C31)
    merge(31, 1, 31, 3)
    c = cell(31, 1)
    c.value = "Superintendent Signature / Firma del superintendente"
    c.font = Font(name="Calibri", size=8)
    c.alignment = center

    c = cell(31, 7)
    c.value = "Date/ Fecha:"
    c.font = Font(name="Calibri", size=11)


def _sanitize_sheet_name(name):
    cleaned = "".join("_" if ch in _INVALID_SHEET_CHARS else ch for ch in name)
    return cleaned.strip()[:31] or "Notice"


def _unique_sheet_name(wb, base):
    name = base[:31]
    if name not in wb.sheetnames:
        return name
    for i in range(2, 1000):
        suffix = f" ({i})"
        candidate = (base[:31 - len(suffix)]) + suffix
        if candidate not in wb.sheetnames:
            return candidate
    return base[:28] + "_X"


def _notice_sheet_name(wb, rec, index, total):
    date_str = ""
    dt = parse_date(rec["date"])
    if dt is not None:
        date_str = dt.strftime("%-m-%-d")
    width = max(2, len(str(total)))
    prefix = f"{index:0{width}d}. "
    budget = 31 - len(prefix)
    body = _sanitize_sheet_name(f"{rec['employee']} {date_str}".strip())[:budget]
    return _unique_sheet_name(wb, prefix + body)


def _apply_page_setup(ws):
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.page_setup.orientation = "landscape"
    ws.page_margins = PageMargins(
        top=0.75, bottom=0.75, left=0.7, right=0.7, header=0.3, footer=0.3
    )


def _write_overview_sheet(ws, records):
    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    title_font = Font(name="Calibri", size=18, bold=True)
    section_font = Font(name="Calibri", size=13, bold=True)
    label_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 26

    def set_row(r, label, value, label_bold=True):
        a = ws.cell(row=r, column=1)
        a.value = label
        a.font = label_font if label_bold else body_font
        a.alignment = left
        b = ws.cell(row=r, column=2)
        b.value = value
        b.font = body_font
        b.alignment = left

    def section_header(r, text, span=2):
        c = ws.cell(row=r, column=1)
        c.value = text
        c.font = section_font
        c.alignment = left
        c.border = border
        for col in range(2, span + 1):
            ws.cell(row=r, column=col).border = border
        if span > 1:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)

    # Title
    t = ws.cell(row=1, column=1)
    t.value = "Overview"
    t.font = title_font
    t.alignment = left

    # Stats
    notices = [r for r in records if r["error"] in TARGET_ERRORS]
    clocked_twice = [r for r in records if r["error"] == "Clocked In Twice"]
    missing_start = [r for r in records if r["error"] == "Missing Start Touch"]
    missing_stop = [r for r in records if r["error"] == "Missing Stop Touch"]

    unique_employees = sorted({r["employee"] for r in records})

    dates = [parse_date(r["date"]) for r in records]
    dates = [d for d in dates if d is not None]
    date_range = ""
    if dates:
        d_min, d_max = min(dates), max(dates)
        if d_min == d_max:
            date_range = d_min.strftime("%m/%d/%Y")
        else:
            date_range = f"{d_min.strftime('%m/%d/%Y')} – {d_max.strftime('%m/%d/%Y')}"

    # Summary block
    set_row(3, "Report Generated:", datetime.now().strftime("%m/%d/%Y %I:%M %p"))
    set_row(4, "Date Range of Incidents:", date_range or "—")
    set_row(5, "Total Records:", len(records))
    set_row(6, "Unique Employees:", len(unique_employees))

    # Notice Sheets Generated
    section_header(8, "Notice Sheets Generated", span=2)

    sheets_header_row = 9
    for col, h in enumerate(["Description", "Count"], start=1):
        c = ws.cell(row=sheets_header_row, column=col)
        c.value = h
        c.font = label_font
        c.alignment = center
        c.border = border

    sheets_rows = [
        ("Total Violations", len(records), False),
        ("Clocked In Twice (no sheet)", len(clocked_twice), False),
        ("Notice Sheets in this File", len(notices), True),
    ]
    for i, (label, count, bold) in enumerate(sheets_rows):
        r = sheets_header_row + 1 + i
        a = ws.cell(row=r, column=1)
        a.value = label
        a.font = label_font if bold else body_font
        a.alignment = left
        a.border = border
        b = ws.cell(row=r, column=2)
        b.value = count
        b.font = label_font if bold else body_font
        b.alignment = center
        b.border = border

    # Breakdown by error type
    breakdown_section_row = sheets_header_row + len(sheets_rows) + 3
    section_header(breakdown_section_row, "Breakdown by Error Type", span=3)

    breakdown_header = breakdown_section_row + 1
    headers = ["Error Type", "Count", "% of Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=breakdown_header, column=col)
        c.value = h
        c.font = label_font
        c.alignment = center
        c.border = border

    total = len(records) or 1
    breakdown_rows = [
        ("Missing Start Touch", len(missing_start)),
        ("Missing Stop Touch", len(missing_stop)),
        ("Clocked In Twice", len(clocked_twice)),
        ("Total", len(records)),
    ]
    for i, (label, count) in enumerate(breakdown_rows):
        r = breakdown_header + 1 + i
        is_total = label == "Total"
        a = ws.cell(row=r, column=1)
        a.value = label
        a.font = label_font if is_total else body_font
        a.alignment = left
        a.border = border
        b = ws.cell(row=r, column=2)
        b.value = count
        b.font = label_font if is_total else body_font
        b.alignment = center
        b.border = border
        c = ws.cell(row=r, column=3)
        c.value = f"{(count / total * 100):.1f}%"
        c.font = label_font if is_total else body_font
        c.alignment = center
        c.border = border

    # Top offenders
    next_row = breakdown_header + len(breakdown_rows) + 3
    section_header(next_row, "Employees with Multiple Violations", span=4)

    emp_counter = Counter(r["employee"] for r in records)
    repeat_offenders = sorted(
        [(name, cnt) for name, cnt in emp_counter.items() if cnt > 1],
        key=lambda x: (-x[1], x[0]),
    )

    emp_header_row = next_row + 1
    headers = ["Employee", "Violations", "Error Types"]
    widths = [1, 2, 4]  # which columns
    for col, h in zip(widths, headers):
        c = ws.cell(row=emp_header_row, column=col)
        c.value = h
        c.font = label_font
        c.alignment = center
        c.border = border

    if not repeat_offenders:
        r = emp_header_row + 1
        c = ws.cell(row=r, column=1)
        c.value = "No employees with more than one violation."
        c.font = body_font
        c.alignment = left
    else:
        for i, (name, cnt) in enumerate(repeat_offenders):
            r = emp_header_row + 1 + i
            types = sorted({rec["error"] for rec in records if rec["employee"] == name})
            a = ws.cell(row=r, column=1)
            a.value = name
            a.font = body_font
            a.alignment = left
            a.border = border
            b = ws.cell(row=r, column=2)
            b.value = cnt
            b.font = body_font
            b.alignment = center
            b.border = border
            # Blank-filled column 3 for grid symmetry
            ws.cell(row=r, column=3).border = border
            d = ws.cell(row=r, column=4)
            d.value = ", ".join(types)
            d.font = body_font
            d.alignment = left
            d.border = border

    # Violations by date
    next_row = emp_header_row + max(len(repeat_offenders), 1) + 3
    section_header(next_row, "Violations by Date", span=2)

    date_header_row = next_row + 1
    for col, h in enumerate(["Date", "Count"], start=1):
        c = ws.cell(row=date_header_row, column=col)
        c.value = h
        c.font = label_font
        c.alignment = center
        c.border = border

    date_counter = Counter()
    for r in records:
        d = parse_date(r["date"])
        if d is not None:
            date_counter[d] += 1

    for i, d in enumerate(sorted(date_counter.keys())):
        r = date_header_row + 1 + i
        a = ws.cell(row=r, column=1)
        a.value = d
        a.number_format = "M/D/YYYY"
        a.font = body_font
        a.alignment = center
        a.border = border
        b = ws.cell(row=r, column=2)
        b.value = date_counter[d]
        b.font = body_font
        b.alignment = center
        b.border = border

    # Clocked In Twice list
    next_row = date_header_row + max(len(date_counter), 1) + 3
    section_header(next_row, "Clocked In Twice", span=2)

    ct_header_row = next_row + 1
    for col, h in enumerate(["Employee", "Date"], start=1):
        c = ws.cell(row=ct_header_row, column=col)
        c.value = h
        c.font = label_font
        c.alignment = center
        c.border = border

    if not clocked_twice:
        r = ct_header_row + 1
        c = ws.cell(row=r, column=1)
        c.value = "No Clocked In Twice records."
        c.font = body_font
        c.alignment = left
    else:
        for i, rec in enumerate(clocked_twice):
            r = ct_header_row + 1 + i
            a = ws.cell(row=r, column=1)
            a.value = rec["employee"]
            a.font = body_font
            a.alignment = left
            a.border = border
            b = ws.cell(row=r, column=2)
            b.value = parse_date(rec["date"])
            b.number_format = "M/D/YYYY"
            b.font = body_font
            b.alignment = center
            b.border = border


def build_workbook(records):
    wb = Workbook()
    # Remove the default sheet created by openpyxl; we'll add our own.
    wb.remove(wb.active)

    notices = [r for r in records if r["error"] in TARGET_ERRORS]

    if records:
        overview_ws = wb.create_sheet("Overview")
        _write_overview_sheet(overview_ws, records)

    for i, rec in enumerate(notices, start=1):
        sheet_name = _notice_sheet_name(wb, rec, i, len(notices))
        ws = wb.create_sheet(sheet_name)
        _apply_page_setup(ws)
        write_notice(ws, 1, rec)
        apply_notice_borders(ws, 1)

    # Safety net: if there were no records at all, make sure the workbook has at least one sheet.
    if not wb.sheetnames:
        wb.create_sheet("Empty")

    return wb
