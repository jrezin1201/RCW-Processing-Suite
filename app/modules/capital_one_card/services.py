"""Capital One Card Report processing service.

Reads a Capital One card export (CSV or XLSX), groups transactions by Category
(blank -> "Uncategorized"), sorts within each category by Description then
Transaction Date, and writes a two-sheet output workbook:

  - ``Analytics`` (active by default) — transaction counts, grand totals
    (live cross-sheet formulas), the card number(s) used, and a per-category
    breakdown of transaction counts and dollar totals.
  - ``Capital One Card`` — the transaction detail with only the columns the
    customer job-costs from (Transaction Date | Description | Debit |
    Credit), a bold subtotal closing each category block, and one bold
    Grand Total at the bottom.

Every transaction stays on its own row — nothing is merged or summarized —
so each purchase can be split and job-costed manually.
"""
from __future__ import annotations

import csv
import io
import logging
from collections import OrderedDict
from collections.abc import Iterator
from datetime import date, datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = ("Transaction Date", "Description", "Category", "Debit", "Credit")
OUTPUT_SHEET_NAME = "Capital One Card"
ANALYTICS_SHEET_NAME = "Analytics"
OUTPUT_HEADERS = ("Transaction Date", "Description", "Debit", "Credit")
GRAND_TOTAL_LABEL = "Grand Total"
UNCATEGORIZED_LABEL = "Uncategorized"

_HEADER_FILL_HEX = "D9D9D9"
_DATA_FONT_NAME = "Arial"
_DATA_FONT_SIZE = 11
_DATE_FORMAT = "m/d/yyyy"
_AMOUNT_FORMAT = "#,##0.00;(#,##0.00)"
_MONEY_FORMAT = "$#,##0.00;($#,##0.00)"
_INT_FORMAT = "#,##0"
_TIMESTAMP_FORMAT = "m/d/yyyy h:mm AM/PM"
_ANALYTICS_TAB_COLOR = "366092"

_DATE_PARSE_FORMATS = ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%y")

_XLSX_MAGIC = b"PK\x03\x04"  # xlsx files are ZIP archives


class CapitalOneCardError(ValueError):
    """Raised when the input file cannot be processed."""


# ---------------------------------------------------------------------------
# loader
# ---------------------------------------------------------------------------

def _norm_header(value) -> str:
    """Lowercase + collapse whitespace so header lookup tolerates minor variations."""
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _cell_text(value) -> str:
    """Render a cell value as trimmed text. Excel often stores numbers like
    card/store numbers as floats — integral floats drop the trailing '.0'."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _coerce_amount(value) -> float | None:
    """Convert a Debit/Credit cell to float; return None for blank/unparseable."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_date(value):
    """Parse a Transaction Date cell to datetime. Excel cells arrive as
    datetime/date objects already; strings are tried against the known
    formats and passed through unchanged if none match (the original value
    is preserved)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()
    if not s:
        return None
    for fmt in _DATE_PARSE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s


def _decode_csv(file_bytes: bytes) -> str:
    """Decode CSV bytes tolerantly (BOM-aware UTF-8, then Latin-1 fallback)."""
    try:
        return file_bytes.decode("utf-8-sig")
    except UnicodeDecodeError:
        return file_bytes.decode("latin-1")


def _rows_from_xlsx(file_bytes: bytes) -> Iterator[list]:
    """Yield rows (lists of raw cell values) from the first sheet of an xlsx."""
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        raise CapitalOneCardError(
            "Could not read the Excel file. Please re-export it and try again."
        ) from exc
    try:
        ws = wb.active
        if ws is not None:
            for row in ws.iter_rows(values_only=True):
                yield list(row)
    finally:
        wb.close()


def _iter_input_rows(file_bytes: bytes) -> Iterator[list]:
    """Dispatch to the xlsx or CSV reader by sniffing the file's magic bytes."""
    if file_bytes[: len(_XLSX_MAGIC)] == _XLSX_MAGIC:
        return _rows_from_xlsx(file_bytes)
    return csv.reader(io.StringIO(_decode_csv(file_bytes)))


def load_transactions(file_bytes: bytes) -> list[dict]:
    """Read transactions from a Capital One export (CSV or XLSX byte stream).

    Validates that every REQUIRED_COLUMNS header is present. Posted Date is
    ignored; Card No. is optional and, when present, captured for the
    Analytics sheet. Blank Category maps to ``Uncategorized``; blank
    Debit/Credit stay ``None`` so their cells render blank while totals treat
    them as 0. Every non-empty row is kept.
    """
    reader = iter(_iter_input_rows(file_bytes))
    try:
        header_row = next(reader)
    except StopIteration as exc:
        raise CapitalOneCardError("Uploaded file is empty.") from exc

    header_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key and key not in header_index:
            header_index[key] = idx

    missing = [c for c in REQUIRED_COLUMNS if _norm_header(c) not in header_index]
    if missing:
        raise CapitalOneCardError(
            f"Missing required column(s): {', '.join(missing)}. "
            "Please upload a Capital One export (CSV or Excel) with Transaction "
            "Date, Description, Category, Debit, and Credit columns."
        )

    date_idx = header_index[_norm_header("Transaction Date")]
    desc_idx = header_index[_norm_header("Description")]
    category_idx = header_index[_norm_header("Category")]
    debit_idx = header_index[_norm_header("Debit")]
    credit_idx = header_index[_norm_header("Credit")]
    card_idx = header_index.get(_norm_header("Card No."))

    def _cell(row: list, idx: int):
        return row[idx] if len(row) > idx else None

    out: list[dict] = []
    for row in reader:
        if not row or all(not _cell_text(c) for c in row):
            continue
        out.append({
            "date": _coerce_date(_cell(row, date_idx)),
            "description": _cell_text(_cell(row, desc_idx)),
            "category": _cell_text(_cell(row, category_idx)) or UNCATEGORIZED_LABEL,
            "debit": _coerce_amount(_cell(row, debit_idx)),
            "credit": _coerce_amount(_cell(row, credit_idx)),
            "card": _cell_text(_cell(row, card_idx)) if card_idx is not None else "",
        })

    if not out:
        raise CapitalOneCardError("No transactions found in uploaded file.")

    return out


# ---------------------------------------------------------------------------
# grouping + sorting
# ---------------------------------------------------------------------------

def group_and_sort(transactions: list[dict]) -> OrderedDict[str, list[dict]]:
    """Group by Category (ascending) and sort each group's rows by Description
    ascending, then Transaction Date ascending. Rows are never merged — each
    transaction remains its own entry."""
    _far_future = datetime.max

    def _date_key(row: dict):
        return row["date"] if isinstance(row["date"], datetime) else _far_future

    groups: dict[str, list[dict]] = {}
    for tx in transactions:
        groups.setdefault(tx["category"], []).append(tx)

    ordered: OrderedDict[str, list[dict]] = OrderedDict()
    for category in sorted(groups.keys(), key=str.casefold):
        ordered[category] = sorted(
            groups[category],
            key=lambda r: (r["description"].casefold(), _date_key(r)),
        )
    return ordered


# ---------------------------------------------------------------------------
# data-sheet writer
# ---------------------------------------------------------------------------

def _build_data_sheet(
    ws: Worksheet,
    grouped: OrderedDict[str, list[dict]],
) -> tuple[OrderedDict[str, int], int]:
    """Populate the transaction sheet and return the cell coordinates the
    Analytics sheet cross-references: an OrderedDict mapping category name to
    its subtotal row, plus the Grand Total row."""
    header_font = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE, bold=True)
    data_font = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE)
    bold_font = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE, bold=True)
    header_fill = PatternFill(start_color=_HEADER_FILL_HEX, end_color=_HEADER_FILL_HEX, fill_type="solid")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_border = Border(top=thin)
    top_bottom_border = Border(top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for col_idx, label in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border

    current_row = 2
    subtotal_row_by_category: OrderedDict[str, int] = OrderedDict()

    for category, rows in grouped.items():
        cat_cell = ws.cell(row=current_row, column=1, value=category)
        cat_cell.font = bold_font
        cat_cell.alignment = left
        for col in range(1, len(OUTPUT_HEADERS) + 1):
            ws.cell(row=current_row, column=col).fill = header_fill
        current_row += 1

        first_row = current_row
        for tx in rows:
            a = ws.cell(row=current_row, column=1, value=tx["date"])
            a.font = data_font
            if isinstance(tx["date"], datetime):
                a.number_format = _DATE_FORMAT
            b = ws.cell(row=current_row, column=2, value=tx["description"])
            b.font = data_font
            for col, amount in ((3, tx["debit"]), (4, tx["credit"])):
                if amount is None:
                    continue
                c = ws.cell(row=current_row, column=col, value=amount)
                c.font = data_font
                c.number_format = _AMOUNT_FORMAT
                c.alignment = right
            current_row += 1
        last_row = current_row - 1

        lbl = ws.cell(row=current_row, column=1, value=f"{category} Total")
        lbl.font = bold_font
        lbl.alignment = left
        lbl.border = top_border
        ws.cell(row=current_row, column=2).border = top_border
        for col_letter, col_idx in (("C", 3), ("D", 4)):
            cell = ws.cell(
                row=current_row,
                column=col_idx,
                value=f"=SUM({col_letter}{first_row}:{col_letter}{last_row})",
            )
            cell.font = bold_font
            cell.number_format = _AMOUNT_FORMAT
            cell.alignment = right
            cell.border = top_border
        subtotal_row_by_category[category] = current_row
        current_row += 1

        current_row += 1  # blank separator between categories

    grand_total_row = current_row
    lbl = ws.cell(row=grand_total_row, column=1, value=GRAND_TOTAL_LABEL)
    lbl.font = bold_font
    lbl.alignment = left
    lbl.border = top_bottom_border
    ws.cell(row=grand_total_row, column=2).border = top_bottom_border
    for col_letter, col_idx in (("C", 3), ("D", 4)):
        # Every transaction belongs to exactly one category block, so summing
        # the category subtotal cells totals each transaction exactly once.
        refs = ",".join(f"{col_letter}{r}" for r in subtotal_row_by_category.values())
        cell = ws.cell(row=grand_total_row, column=col_idx, value=f"=SUM({refs})")
        cell.font = bold_font
        cell.number_format = _AMOUNT_FORMAT
        cell.alignment = right
        cell.border = top_bottom_border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A2"

    return subtotal_row_by_category, grand_total_row


# ---------------------------------------------------------------------------
# Analytics-sheet writer
# ---------------------------------------------------------------------------

def _format_date_short(d: datetime) -> str:
    """Render a date as ``M/D/YYYY`` portably (no platform-specific strftime)."""
    return f"{d.month}/{d.day}/{d.year}"


def _build_analytics_sheet(
    ws: Worksheet,
    grouped: OrderedDict[str, list[dict]],
    subtotal_row_by_category: OrderedDict[str, int],
    grand_total_row: int,
    generated_at: datetime,
) -> None:
    """Populate the Analytics dashboard sheet."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _ANALYTICS_TAB_COLOR

    arial = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE)
    arial_bold = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE, bold=True)
    title_font = Font(name=_DATA_FONT_NAME, size=16, bold=True)
    section_fill = PatternFill(start_color=_HEADER_FILL_HEX, end_color=_HEADER_FILL_HEX, fill_type="solid")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    label_col = 2

    title = ws.cell(row=2, column=label_col, value="Capital One Card Report")
    title.font = title_font
    title.alignment = left

    def _kv(row: int, label: str, value, *, bold=True, fmt=None) -> None:
        lbl = ws.cell(row=row, column=label_col, value=label)
        lbl.font = arial
        lbl.alignment = right
        v = ws.cell(row=row, column=label_col + 1, value=value)
        v.font = arial_bold if bold else arial
        v.alignment = left
        if fmt:
            v.number_format = fmt

    def _section_header(row: int, label: str, *, span: int = 2) -> None:
        for col in range(label_col, label_col + span):
            ws.cell(row=row, column=col).fill = section_fill
        c = ws.cell(row=row, column=label_col, value=label)
        c.font = arial_bold
        c.alignment = left

    all_txs = [t for rows in grouped.values() for t in rows]

    # Row 4: Generated timestamp.
    _kv(4, "Generated:", generated_at, fmt=_TIMESTAMP_FORMAT)

    # Row 5: Date range from input.
    tx_dates = [t["date"] for t in all_txs if isinstance(t["date"], datetime)]
    if tx_dates:
        d0, d1 = min(tx_dates), max(tx_dates)
        date_range = f"{_format_date_short(d0)} – {_format_date_short(d1)}"
    else:
        date_range = ""
    _kv(5, "Date range (Transaction):", date_range)

    # Row 6: Card number(s) used.
    cards = sorted({t["card"] for t in all_txs if t["card"]})
    _kv(6, "Card number(s):", ", ".join(cards) if cards else "Not in input file")

    def _formula_to(col_letter: str, row: int) -> str:
        return f"='{OUTPUT_SHEET_NAME}'!{col_letter}{row}"

    row = 8
    _section_header(row, "OVERVIEW")
    row += 1
    _kv(row, "Total transactions:", len(all_txs), fmt=_INT_FORMAT)
    row += 1
    _kv(row, "Grand Total (Debit):", _formula_to("C", grand_total_row), fmt=_MONEY_FORMAT)
    row += 1
    _kv(row, "Grand Total (Credit):", _formula_to("D", grand_total_row), fmt=_MONEY_FORMAT)
    row += 1

    row += 1  # spacer
    _section_header(row, "BY CATEGORY", span=4)
    row += 1

    for col_offset, header in enumerate(("Category", "Transactions", "Debit", "Credit")):
        c = ws.cell(row=row, column=label_col + col_offset, value=header)
        c.font = arial_bold
        c.alignment = left if col_offset == 0 else right
    row += 1

    for category, rows in grouped.items():
        subtotal_row = subtotal_row_by_category[category]
        name = ws.cell(row=row, column=label_col, value=category)
        name.font = arial
        name.alignment = left
        count = ws.cell(row=row, column=label_col + 1, value=len(rows))
        count.font = arial
        count.alignment = right
        count.number_format = _INT_FORMAT
        for col_offset, col_letter in ((2, "C"), (3, "D")):
            c = ws.cell(row=row, column=label_col + col_offset, value=_formula_to(col_letter, subtotal_row))
            c.font = arial
            c.alignment = right
            c.number_format = _MONEY_FORMAT
        row += 1


# ---------------------------------------------------------------------------
# top-level writer + orchestration
# ---------------------------------------------------------------------------

def write_report(
    grouped: OrderedDict[str, list[dict]],
    *,
    generated_at: datetime | None = None,
) -> bytes:
    """Build the two-sheet output workbook and return its bytes.

    The Analytics sheet is built second (so it can reference the freshly
    populated data cells) but moved to index 0 and set as the active sheet,
    so it opens by default in Excel — same convention as Merchant Charges.
    """
    if generated_at is None:
        generated_at = datetime.now()

    wb = openpyxl.Workbook()
    data_ws = wb.active
    data_ws.title = OUTPUT_SHEET_NAME

    subtotal_row_by_category, grand_total_row = _build_data_sheet(data_ws, grouped)

    analytics_ws = wb.create_sheet(ANALYTICS_SHEET_NAME)
    _build_analytics_sheet(
        analytics_ws, grouped, subtotal_row_by_category, grand_total_row, generated_at
    )

    analytics_idx = wb.sheetnames.index(ANALYTICS_SHEET_NAME)
    if analytics_idx != 0:
        wb.move_sheet(ANALYTICS_SHEET_NAME, offset=-analytics_idx)
    wb.active = wb.sheetnames.index(ANALYTICS_SHEET_NAME)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def process_capital_one_card(file_bytes: bytes) -> bytes:
    """Top-level orchestration: validate → load → group + sort → write."""
    transactions = load_transactions(file_bytes)
    grouped = group_and_sort(transactions)

    # Accounting identity: grouping only rearranges rows, never drops or merges.
    grouped_count = sum(len(rows) for rows in grouped.values())
    assert grouped_count == len(transactions), (
        f"capital_one_card accounting mismatch: grouped={grouped_count} "
        f"input={len(transactions)}"
    )

    return write_report(grouped)


__all__ = [
    "CapitalOneCardError",
    "REQUIRED_COLUMNS",
    "OUTPUT_SHEET_NAME",
    "ANALYTICS_SHEET_NAME",
    "OUTPUT_HEADERS",
    "GRAND_TOTAL_LABEL",
    "UNCATEGORIZED_LABEL",
    "load_transactions",
    "group_and_sort",
    "write_report",
    "process_capital_one_card",
]
