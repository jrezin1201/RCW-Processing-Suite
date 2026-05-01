"""Merchant Charges processing service.

Reads a credit-card transactions XLSX, groups positive transactions by brand
(store-number suffixes stripped), sorts within each brand, and writes a
two-sheet output workbook:

  - ``Summary`` (active by default) — a compact dashboard with transaction
    counts, the date range, dollar totals (live cross-sheet formulas), and
    the brand-group count.
  - ``Merchant Charges`` — the detailed transaction data with per-brand
    subtotals, a Credits & Refunds section for negative transactions, and
    dual grand totals (Charges Only + Net) — all as live formulas.
"""
from __future__ import annotations

import logging
import re
from collections import Counter, OrderedDict
from datetime import datetime
from io import BytesIO
from typing import NamedTuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

REQUIRED_COLUMNS = ("Posted", "Occurred", "Merchant Name", "Billed Amount")
OUTPUT_SHEET_NAME = "Merchant Charges"
SUMMARY_SHEET_NAME = "Summary"
OUTPUT_HEADERS = ("Posted Date", "Occurred Date", "Merchant Name", "Amount", "Merchant Total")
CREDITS_SECTION_LABEL = "Credits & Refunds"

DROP_REASON_BLANK_MERCHANT = "blank merchant name"
DROP_REASON_BAD_AMOUNT = "unparseable amount"

_HEADER_FILL_HEX = "D9D9D9"
_DATA_FONT_NAME = "Arial"
_DATA_FONT_SIZE = 11
_DATE_FORMAT = "m/d/yyyy"
_AMOUNT_FORMAT = "#,##0.00;(#,##0.00)"
_MONEY_FORMAT = "$#,##0.00;($#,##0.00)"
_INT_FORMAT = "#,##0"
_TIMESTAMP_FORMAT = "m/d/yyyy h:mm AM/PM"
_SUMMARY_TAB_COLOR = "366092"

# Strips a trailing store-number suffix from a merchant string, e.g.
# ' #0668', ' 912987', '114' (no separator), '#01574*', ' # 170'. Applied
# repeatedly until no further match so multi-token tails collapse cleanly.
_TRAILING_STORE_NUM_RE = re.compile(r"\s*#?\s*\d+\*?\s*$")


class MerchantChargesError(ValueError):
    """Raised when the input workbook cannot be processed."""


class LoadStats(NamedTuple):
    """Per-load statistics used to populate the Summary sheet.

    ``total_input_rows`` is every non-empty data row encountered (excluding
    the header). Surviving rows partition into positive + negative; dropped
    rows are accounted for in ``drop_counts``. The accounting identity
    ``positive + negative + sum(drop_counts.values()) == total_input_rows``
    is asserted in :func:`process_merchant_charges`.
    """
    total_input_rows: int
    drop_counts: dict[str, int]


# ---------------------------------------------------------------------------
# loader
# ---------------------------------------------------------------------------

def _norm_header(value) -> str:
    """Lowercase + collapse whitespace so header lookup tolerates minor variations."""
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def _coerce_amount(value) -> float | None:
    """Convert a Billed Amount cell to float; return None if it cannot be parsed."""
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = s.replace("$", "").replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = "-" + s[1:-1]
    try:
        return float(s)
    except ValueError:
        return None


def _coerce_date(value):
    """Return a datetime if value is date-like; otherwise pass through unchanged."""
    if value is None or isinstance(value, datetime):
        return value
    return value


def load_transactions_with_stats(file_bytes: bytes) -> tuple[list[dict], LoadStats]:
    """Read transactions from an XLSX byte stream and return ``(transactions, stats)``.

    ``stats.total_input_rows`` counts every non-empty data row encountered
    (i.e. each row that the loader inspected). ``stats.drop_counts`` is a
    mapping from drop reason to count, populated only for reasons that
    actually applied to this file.
    """
    try:
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise MerchantChargesError(f"Unable to read workbook: {exc}") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
    except StopIteration as exc:
        raise MerchantChargesError("Workbook is empty.") from exc

    header_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key and key not in header_index:
            header_index[key] = idx

    missing = [c for c in REQUIRED_COLUMNS if _norm_header(c) not in header_index]
    if missing:
        raise MerchantChargesError(
            f"Missing required column(s): {', '.join(missing)}"
        )

    posted_idx = header_index[_norm_header("Posted")]
    occurred_idx = header_index[_norm_header("Occurred")]
    merchant_idx = header_index[_norm_header("Merchant Name")]
    amount_idx = header_index[_norm_header("Billed Amount")]

    out: list[dict] = []
    total_input_rows = 0
    dropped_blank_merchant = 0
    dropped_bad_amount = 0

    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        total_input_rows += 1
        merchant_raw = row[merchant_idx] if len(row) > merchant_idx else None
        merchant = str(merchant_raw).strip() if merchant_raw is not None else ""
        if not merchant:
            dropped_blank_merchant += 1
            continue

        amount = _coerce_amount(row[amount_idx] if len(row) > amount_idx else None)
        if amount is None:
            dropped_bad_amount += 1
            continue

        posted = _coerce_date(row[posted_idx] if len(row) > posted_idx else None)
        occurred = _coerce_date(row[occurred_idx] if len(row) > occurred_idx else None)

        out.append({
            "posted": posted,
            "occurred": occurred,
            "merchant": merchant,
            "amount": amount,
        })

    if dropped_blank_merchant:
        logger.info("merchant_charges: dropped %d rows with blank Merchant Name", dropped_blank_merchant)
    if dropped_bad_amount:
        logger.info("merchant_charges: dropped %d rows with unparseable Billed Amount", dropped_bad_amount)

    if not out:
        raise MerchantChargesError("No valid transactions found in uploaded file.")

    drop_counts: dict[str, int] = {}
    if dropped_blank_merchant:
        drop_counts[DROP_REASON_BLANK_MERCHANT] = dropped_blank_merchant
    if dropped_bad_amount:
        drop_counts[DROP_REASON_BAD_AMOUNT] = dropped_bad_amount

    return out, LoadStats(total_input_rows=total_input_rows, drop_counts=drop_counts)


def load_transactions(file_bytes: bytes) -> list[dict]:
    """Convenience wrapper returning just the cleaned transactions list."""
    txs, _ = load_transactions_with_stats(file_bytes)
    return txs


# ---------------------------------------------------------------------------
# grouping + sorting
# ---------------------------------------------------------------------------

def _merchant_sort_key(name: str) -> tuple[str, str]:
    """Sort merchants like Excel: alphanumerics first (case-insensitive), punctuation
    treated as a tiebreaker. So '76 - TEMECULA' precedes '7-ELEVEN' because the '6'
    in '76' is compared before the '-' in '7-ELEVEN' once spaces/dashes are ignored.
    """
    primary = "".join(c for c in name.casefold() if c.isalnum())
    return (primary, name.casefold())


def _brand_of(merchant: str) -> str:
    """Collapse a full merchant string to its brand by stripping trailing
    store-number suffixes and normalizing internal whitespace. Examples:
        'THE HOME DEPOT #0668' -> 'THE HOME DEPOT'
        'THE HOME DEPOT 673'   -> 'THE HOME DEPOT'
        'STATERBROS114'        -> 'STATERBROS'
        'LOWES #01574*'        -> 'LOWES'
        'DUNN-EDWARDS # 170'   -> 'DUNN-EDWARDS'
        'DUNN-EDWARDS CORP #146' -> 'DUNN-EDWARDS CORP'  (CORP is alpha — kept)
        '76 - TEMECULA FUELS CORP' -> '76 - TEMECULA FUELS CORP'  (no trailing digits)
        'BURGER  KING 09812'   -> 'BURGER KING'  (double space collapsed)
    """
    s = merchant.strip()
    while True:
        nxt = _TRAILING_STORE_NUM_RE.sub("", s)
        if nxt == s:
            break
        s = nxt
    # Collapse runs of internal whitespace to single spaces so banks that
    # double-space "BURGER  KING" produce the same normalized form as a
    # cleanly-spaced "BURGER KING".
    return " ".join(s.split())


def _to_match_key(normalized_brand: str) -> str:
    """Whitespace-insensitive grouping key: removes ALL internal whitespace,
    preserves every other character. Used only for grouping decisions, not
    for display. ``BURGER KING`` and ``BURGERKING`` both yield ``BURGERKING``
    and merge into the same group; ``WAL-MART`` and ``WALMART`` differ in
    punctuation so their match keys differ and they stay separate.
    """
    return "".join(normalized_brand.split())


def _pick_display_brand(normalized_forms: list[str]) -> str:
    """Choose the display label for a merged brand group:
      1. Prefer forms that contain internal whitespace (so ``BURGER KING`` wins
         over ``BURGERKING``).
      2. Among the candidates from rule 1, the most frequent in the input.
      3. Alphabetical ascending tiebreaker.
    Rule 1 is intentionally absolute: a single spaced occurrence beats five
    unspaced ones.
    """
    with_ws = [n for n in normalized_forms if " " in n]
    candidates = with_ws if with_ws else normalized_forms
    counts = Counter(candidates)
    max_count = max(counts.values())
    most_common = [n for n, c in counts.items() if c == max_count]
    return sorted(most_common)[0]


def group_and_sort(
    transactions: list[dict],
) -> tuple[OrderedDict[str, list[dict]], list[dict]]:
    """Partition transactions by sign and group positives by brand.

    Grouping is whitespace-insensitive: two normalized brand strings collapse
    into the same group when they are equal once all internal whitespace is
    removed. ``BURGER KING`` and ``BURGERKING`` therefore merge under the
    match key ``BURGERKING``, and the displayed group label is the form with
    whitespace (``BURGER KING``) per :func:`_pick_display_brand`. Hyphens,
    punctuation, and abbreviations are NOT collapsed — only whitespace.

    Returns ``(positive_groups, negative_rows)`` where:
      - ``positive_groups`` is an OrderedDict keyed by display label
        (alphabetical via ``_merchant_sort_key`` applied to the label); each
        value is the merged group's transaction rows sorted by full original
        merchant name, then Posted date asc, then Occurred date asc.
      - ``negative_rows`` is a flat list of every transaction with
        ``amount < 0``, sorted by Posted asc, then Occurred asc, then merchant
        name asc.

    A brand whose only transactions are negative does not appear in
    ``positive_groups`` at all — every negative row goes to the Credits &
    Refunds section regardless of its brand.
    """
    positives: list[dict] = []
    negatives: list[dict] = []
    for tx in transactions:
        if tx["amount"] < 0:
            negatives.append(tx)
        else:
            positives.append(tx)

    # Group by whitespace-stripped match key. Keep the per-row normalized
    # form alongside so the display-label picker has the full multiset.
    rows_by_key: dict[str, list[dict]] = {}
    forms_by_key: dict[str, list[str]] = {}
    for tx in positives:
        normalized = _brand_of(tx["merchant"])
        match_key = _to_match_key(normalized)
        rows_by_key.setdefault(match_key, []).append(tx)
        forms_by_key.setdefault(match_key, []).append(normalized)

    display_by_key = {k: _pick_display_brand(forms_by_key[k]) for k in rows_by_key}

    _far_future = datetime.max
    ordered: OrderedDict[str, list[dict]] = OrderedDict()
    for match_key in sorted(rows_by_key.keys(), key=lambda k: _merchant_sort_key(display_by_key[k])):
        rows = sorted(
            rows_by_key[match_key],
            key=lambda r: (
                _merchant_sort_key(r["merchant"]),
                r["posted"] or _far_future,
                r["occurred"] or _far_future,
            ),
        )
        ordered[display_by_key[match_key]] = rows

    negatives.sort(
        key=lambda r: (
            r["posted"] or _far_future,
            r["occurred"] or _far_future,
            r["merchant"].casefold(),
        )
    )

    return ordered, negatives


# ---------------------------------------------------------------------------
# data-sheet writer (was ``write_report``; now produces just the data sheet
# and returns the cell addresses the Summary sheet needs to reference)
# ---------------------------------------------------------------------------

class _DataSheetAddresses(NamedTuple):
    """Cell addresses on the Merchant Charges sheet that the Summary sheet
    cross-references. ``credits_subtotal`` and ``net`` are ``None`` when
    there are no negative transactions."""
    charges_only: str | None
    credits_subtotal: str | None
    net: str | None
    brand_group_count: int


def _build_data_sheet(
    ws: Worksheet,
    grouped: OrderedDict[str, list[dict]],
    negatives: list[dict],
) -> _DataSheetAddresses:
    """Populate the Merchant Charges sheet and return key cell addresses."""
    header_font = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE, bold=True)
    data_font = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE)
    bold_font = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE, bold=True)
    header_fill = PatternFill(start_color=_HEADER_FILL_HEX, end_color=_HEADER_FILL_HEX, fill_type="solid")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    top_border = Border(top=thin)
    top_bottom_border = Border(top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    for col_idx, label in enumerate(OUTPUT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = header_border

    def _write_data_row(row: int, tx: dict) -> None:
        a = ws.cell(row=row, column=1, value=tx["posted"])
        a.font = data_font
        a.number_format = _DATE_FORMAT
        b = ws.cell(row=row, column=2, value=tx["occurred"])
        b.font = data_font
        b.number_format = _DATE_FORMAT
        c = ws.cell(row=row, column=3, value=tx["merchant"])
        c.font = data_font
        d = ws.cell(row=row, column=4, value=tx["amount"])
        d.font = data_font
        d.number_format = _AMOUNT_FORMAT
        d.alignment = right

    def _write_subtotal(row: int, formula: str) -> None:
        for col in (4, 5):
            cell = ws.cell(row=row, column=col, value=formula)
            cell.font = bold_font
            cell.number_format = _AMOUNT_FORMAT
            cell.alignment = right
            cell.border = top_border

    brand_subtotal_cells: list[str] = []
    current_row = 2

    for _brand, rows in grouped.items():
        first_row = current_row
        for tx in rows:
            _write_data_row(current_row, tx)
            current_row += 1
        last_row = current_row - 1
        _write_subtotal(current_row, f"=SUM(D{first_row}:D{last_row})")
        brand_subtotal_cells.append(f"E{current_row}")
        current_row += 1  # subtotal row consumed
        current_row += 1  # blank separator

    credits_subtotal_cell: str | None = None
    if negatives:
        hdr = ws.cell(row=current_row, column=3, value=CREDITS_SECTION_LABEL)
        hdr.font = bold_font
        hdr.fill = header_fill
        hdr.alignment = left
        current_row += 1

        first_neg_row = current_row
        for tx in negatives:
            _write_data_row(current_row, tx)
            current_row += 1
        last_neg_row = current_row - 1

        _write_subtotal(current_row, f"=SUM(D{first_neg_row}:D{last_neg_row})")
        credits_subtotal_cell = f"E{current_row}"
        current_row += 1

    current_row += 1  # blank row before grand totals

    has_brands = bool(brand_subtotal_cells)
    has_credits = credits_subtotal_cell is not None

    def _write_grand_total(row: int, label: str, formula: str, border) -> str:
        c = ws.cell(row=row, column=3, value=label)
        c.font = bold_font
        c.alignment = right
        e = ws.cell(row=row, column=5, value=formula)
        e.font = bold_font
        e.number_format = _AMOUNT_FORMAT
        e.alignment = right
        e.border = border
        return f"E{row}"

    charges_only_cell: str | None = None
    net_cell: str | None = None

    if has_brands and has_credits:
        charges_formula = "=SUM(" + ",".join(brand_subtotal_cells) + ")"
        charges_only_cell = _write_grand_total(
            current_row, "Grand Total (Charges Only)", charges_formula, top_border
        )
        charges_row = current_row
        current_row += 1
        net_cell = _write_grand_total(
            current_row,
            "Grand Total (Net)",
            f"=E{charges_row}+{credits_subtotal_cell}",
            top_bottom_border,
        )
    elif has_brands:
        formula = "=SUM(" + ",".join(brand_subtotal_cells) + ")"
        charges_only_cell = _write_grand_total(
            current_row, "Grand Total", formula, top_bottom_border
        )
    elif has_credits:
        # All-negative input: render a single Grand Total (Net) referencing
        # the credits subtotal directly. Charges Only is omitted because
        # there are no charges to report.
        net_cell = _write_grand_total(
            current_row,
            "Grand Total (Net)",
            f"={credits_subtotal_cell}",
            top_bottom_border,
        )

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.freeze_panes = "A2"

    return _DataSheetAddresses(
        charges_only=charges_only_cell,
        credits_subtotal=credits_subtotal_cell,
        net=net_cell,
        brand_group_count=len(grouped),
    )


# ---------------------------------------------------------------------------
# Summary-sheet writer
# ---------------------------------------------------------------------------

def _format_date_short(d: datetime) -> str:
    """Render a date as ``M/D/YYYY`` portably (no platform-specific strftime)."""
    return f"{d.month}/{d.day}/{d.year}"


def _build_summary_sheet(
    ws: Worksheet,
    grouped: OrderedDict[str, list[dict]],
    negatives: list[dict],
    stats: LoadStats,
    addrs: _DataSheetAddresses,
    generated_at: datetime,
) -> None:
    """Populate the Summary dashboard sheet."""
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = _SUMMARY_TAB_COLOR

    arial = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE)
    arial_bold = Font(name=_DATA_FONT_NAME, size=_DATA_FONT_SIZE, bold=True)
    title_font = Font(name=_DATA_FONT_NAME, size=16, bold=True)
    section_fill = PatternFill(start_color=_HEADER_FILL_HEX, end_color=_HEADER_FILL_HEX, fill_type="solid")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    top_border = Border(top=Side(style="thin", color="000000"))

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22

    label_col = 2
    value_col = 3

    title = ws.cell(row=2, column=label_col, value="Merchant Charges Report")
    title.font = title_font
    title.alignment = left

    def _kv(row: int, label: str, value, *, bold=True, fmt=None, border=None) -> None:
        lbl = ws.cell(row=row, column=label_col, value=label)
        lbl.font = arial
        lbl.alignment = right
        v = ws.cell(row=row, column=value_col, value=value)
        v.font = arial_bold if bold else arial
        v.alignment = left
        if fmt:
            v.number_format = fmt
        if border:
            v.border = border

    def _section_header(row: int, label: str) -> None:
        for col in (label_col, value_col):
            ws.cell(row=row, column=col).fill = section_fill
        c = ws.cell(row=row, column=label_col, value=label)
        c.font = arial_bold
        c.alignment = left

    # Row 4: Generated timestamp.
    _kv(4, "Generated:", generated_at, fmt=_TIMESTAMP_FORMAT)

    # Row 5: Date range from input.
    all_txs = [t for rs in grouped.values() for t in rs] + negatives
    posted_dates = [t["posted"] for t in all_txs if t["posted"] is not None]
    if posted_dates:
        d0, d1 = min(posted_dates), max(posted_dates)
        date_range = f"{_format_date_short(d0)} \u2013 {_format_date_short(d1)}"
    else:
        date_range = ""
    _kv(5, "Date range (Posted):", date_range)

    row = 7
    _section_header(row, "TRANSACTION COUNTS")
    row += 1

    pos_count = sum(len(rs) for rs in grouped.values())
    neg_count = len(negatives)
    drop_total = sum(stats.drop_counts.values())

    _kv(row, "Total transactions in input:", stats.total_input_rows, fmt=_INT_FORMAT)
    row += 1
    _kv(row, "Charges (positive):", pos_count, fmt=_INT_FORMAT)
    row += 1
    _kv(row, "Credits (negative):", neg_count, fmt=_INT_FORMAT)
    row += 1
    if drop_total > 0:
        _kv(row, "Rows dropped (validation):", drop_total, fmt=_INT_FORMAT)
        row += 1
        for reason in sorted(stats.drop_counts.keys()):
            _kv(row, f"   {reason}:", stats.drop_counts[reason], bold=False, fmt=_INT_FORMAT)
            row += 1

    row += 1  # spacer
    _section_header(row, "DOLLAR TOTALS")
    row += 1

    def _formula_to(cell: str) -> str:
        return f"='{OUTPUT_SHEET_NAME}'!{cell}"

    if addrs.charges_only is not None:
        _kv(row, "Grand Total (Charges Only):", _formula_to(addrs.charges_only), fmt=_MONEY_FORMAT)
        row += 1
    if neg_count > 0 and addrs.credits_subtotal is not None:
        _kv(row, "Credits & Refunds subtotal:", _formula_to(addrs.credits_subtotal), fmt=_MONEY_FORMAT)
        row += 1
    if neg_count > 0 and addrs.net is not None:
        _kv(row, "Grand Total (Net):", _formula_to(addrs.net), fmt=_MONEY_FORMAT, border=top_border)
        row += 1

    row += 1  # spacer
    _section_header(row, "BRAND GROUPS")
    row += 1
    _kv(row, "Total brand groups:", addrs.brand_group_count, fmt=_INT_FORMAT)


# ---------------------------------------------------------------------------
# top-level writer + orchestration
# ---------------------------------------------------------------------------

def write_report(
    grouped: OrderedDict[str, list[dict]],
    negatives: list[dict] | None = None,
    *,
    stats: LoadStats | None = None,
    generated_at: datetime | None = None,
) -> bytes:
    """Build the two-sheet output workbook and return its bytes.

    The Summary sheet is built second (so it can reference the freshly
    populated Merchant Charges cells) but moved to index 0 and set as the
    active sheet, so it opens by default in Excel.

    ``stats`` is optional: when omitted, a synthetic LoadStats is computed
    from the partitioned data with no validation drops. The orchestration
    in :func:`process_merchant_charges` always passes a real LoadStats.
    """
    if negatives is None:
        negatives = []
    if generated_at is None:
        generated_at = datetime.now()

    pos_count = sum(len(rs) for rs in grouped.values())
    neg_count = len(negatives)
    if stats is None:
        stats = LoadStats(total_input_rows=pos_count + neg_count, drop_counts={})

    wb = openpyxl.Workbook()
    data_ws = wb.active
    data_ws.title = OUTPUT_SHEET_NAME

    addrs = _build_data_sheet(data_ws, grouped, negatives)

    summary_ws = wb.create_sheet(SUMMARY_SHEET_NAME)
    _build_summary_sheet(summary_ws, grouped, negatives, stats, addrs, generated_at)

    # Move Summary to index 0 and make it the active tab so it opens by default.
    summary_idx = wb.sheetnames.index(SUMMARY_SHEET_NAME)
    if summary_idx != 0:
        wb.move_sheet(SUMMARY_SHEET_NAME, offset=-summary_idx)
    wb.active = wb.sheetnames.index(SUMMARY_SHEET_NAME)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def process_merchant_charges(file_bytes: bytes) -> bytes:
    """Top-level orchestration: validate → load → partition + group → write."""
    transactions, stats = load_transactions_with_stats(file_bytes)
    positive_groups, negatives = group_and_sort(transactions)

    pos_count = sum(len(rs) for rs in positive_groups.values())
    neg_count = len(negatives)
    drop_total = sum(stats.drop_counts.values())
    # Accounting identity: every input row is either positive, negative, or dropped.
    assert pos_count + neg_count + drop_total == stats.total_input_rows, (
        f"merchant_charges accounting mismatch: positive={pos_count} "
        f"negative={neg_count} dropped={drop_total} input={stats.total_input_rows}"
    )

    return write_report(positive_groups, negatives, stats=stats)


# Re-export utilities used by the column-letter logic in tests.
__all__ = [
    "MerchantChargesError",
    "LoadStats",
    "OUTPUT_SHEET_NAME",
    "SUMMARY_SHEET_NAME",
    "OUTPUT_HEADERS",
    "CREDITS_SECTION_LABEL",
    "DROP_REASON_BLANK_MERCHANT",
    "DROP_REASON_BAD_AMOUNT",
    "load_transactions",
    "load_transactions_with_stats",
    "group_and_sort",
    "write_report",
    "process_merchant_charges",
    "_brand_of",
    "_merchant_sort_key",
    "get_column_letter",
]
