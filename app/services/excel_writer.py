"""Excel output writer service for generating formatted summary files."""
import os
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models.schemas import SummaryRow, QAReport


def write_summary_excel(
    summary_rows: List[Dict[str, Any]],
    qa_report: QAReport,
    job_id: str,
    category_headers: List[str],
    phase: str = None,
    project_name: str = None,
    house_string: str = None,
    original_filename: str = None
) -> str:
    """
    Write summary data to a formatted Excel file with dynamic category columns.

    Args:
        summary_rows: List of aggregated summary rows (as dicts)
        qa_report: QA report data
        job_id: Job identifier for filename
        category_headers: List of category column headers (in order)
        phase: Phase number extracted from input file
        project_name: Project name extracted from B3
        house_string: House string extracted from B5
        original_filename: Original uploaded filename (without extension)

    Returns:
        Path to the generated Excel file
    """
    # Create output directory if it doesn't exist
    output_dir = Path("data/outputs")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Create workbook and worksheets
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Define styles
    header_font = Font(bold=True, size=16)
    normal_font = Font(size=16)
    accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_blue_fill = PatternFill(start_color="E0F2F7", end_color="E0F2F7", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Row 1: Project header
    ws["B1"] = f"Project Name: {project_name if project_name else ''}"
    ws["G1"] = f"Phase:{phase if phase else ''}"
    ws["H1"] = house_string if house_string else ""
    ws["I1"] = "Job #:"

    # Merge cells B1, C1, and D1 for Project Name
    ws.merge_cells('B1:D1')

    ws["B1"].font = header_font
    ws["G1"].font = header_font
    ws["H1"].font = header_font
    ws["H1"].fill = light_blue_fill
    ws["I1"].font = header_font

    # Row 2: Empty

    # Row 3: Headers
    # Build headers: LOT, PLAN, [category columns], Total
    all_headers = ["LOT", "PLAN"] + category_headers + ["Total"]
    header_row = 3

    # Column A3 is blank
    ws.cell(row=header_row, column=1, value="")

    for col_idx, header in enumerate(all_headers, 2):  # Start from column 2 (B)
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if header == "Total":
            cell.fill = yellow_fill
        else:
            cell.fill = light_gray_fill

    # Calculate total column index
    total_col_idx = len(all_headers) + 1  # +1 because we start from column 2

    # Write data rows starting from row 4
    data_start_row = 4
    for row_num, summary_row in enumerate(summary_rows, 1):
        row_idx = data_start_row + row_num - 1

        # Column A: Row number
        ws.cell(row=row_idx, column=1, value=row_num).font = normal_font

        # Column B: LOT
        lot_value = summary_row.get("lot_block", "")
        try:
            lot_value = int(lot_value)
        except (ValueError, TypeError):
            try:
                lot_value = float(lot_value)
            except (ValueError, TypeError):
                pass
        ws.cell(row=row_idx, column=2, value=lot_value).font = normal_font

        # Column C: PLAN
        ws.cell(row=row_idx, column=3, value=summary_row.get("plan", "")).font = normal_font

        # Category columns (starting from column D)
        for cat_idx, header in enumerate(category_headers):
            col = 4 + cat_idx  # Column D is column 4
            value = summary_row.get(header, 0.0)
            cell = ws.cell(row=row_idx, column=col, value=value if value else 0)
            cell.number_format = accounting_format
            cell.font = normal_font

        # Total column (as SUM formula)
        first_cat_col = get_column_letter(4)  # Column D
        last_cat_col = get_column_letter(4 + len(category_headers) - 1)
        total_col = 4 + len(category_headers)
        total_formula = f"=SUM({first_cat_col}{row_idx}:{last_cat_col}{row_idx})"
        total_cell = ws.cell(row=row_idx, column=total_col, value=total_formula)
        total_cell.number_format = accounting_format
        total_cell.font = normal_font
        total_cell.fill = yellow_fill

    # After last data row: Skip one row (empty row)
    total_row = data_start_row + len(summary_rows) + 1

    # Apply yellow fill to Total column from header through totals
    total_col = 4 + len(category_headers)
    for row in range(header_row, max(total_row + 1, 14)):
        cell = ws.cell(row=row, column=total_col)
        cell.fill = yellow_fill

    # Create larger font for total row
    total_font = Font(bold=True, size=16)

    # Add "TOTAL" label (in the column before Total)
    label_col = total_col - 1
    total_label_cell = ws.cell(row=total_row, column=label_col, value="TOTAL")
    total_label_cell.font = total_font

    # Calculate and write sum in Total column
    col_letter = get_column_letter(total_col)
    formula = f"=SUM({col_letter}{data_start_row}:{col_letter}{total_row-2})"
    cell = ws.cell(row=total_row, column=total_col, value=formula)
    cell.number_format = accounting_format
    cell.font = total_font
    cell.fill = yellow_fill

    # Skip 2 rows after total, then add Labor row
    labor_row = total_row + 3
    material_row = labor_row + 2

    # Add "LABOR:" in column D
    labor_cell = ws.cell(row=labor_row, column=4, value="LABOR:")
    labor_cell.font = normal_font

    # Add labor calculation in column E (43% of total)
    labor_calc_cell = ws.cell(row=labor_row, column=5, value=f"={col_letter}{total_row}*0.43")
    labor_calc_cell.number_format = accounting_format
    labor_calc_cell.font = normal_font

    # Add description in column G for labor
    desc_labor = ws.cell(row=labor_row, column=7, value="Will be 43% of total amount")
    desc_labor.font = normal_font

    # Add "MATERIAL:" in column D
    material_cell = ws.cell(row=material_row, column=4, value="MATERIAL:")
    material_cell.font = normal_font

    # Add material calculation in column E (28% of total)
    material_calc_cell = ws.cell(row=material_row, column=5, value=f"={col_letter}{total_row}*0.28")
    material_calc_cell.number_format = accounting_format
    material_calc_cell.font = normal_font

    # Add description in column G for material
    desc_material = ws.cell(row=material_row, column=7, value="will be 28% of total amount")
    desc_material.font = normal_font

    # Set column widths
    ws.column_dimensions["A"].width = 4     # Row number
    ws.column_dimensions["B"].width = 7     # LOT
    ws.column_dimensions["C"].width = 7     # PLAN

    # Set widths for category columns
    for cat_idx in range(len(category_headers)):
        col_letter = get_column_letter(4 + cat_idx)
        # Wider for longer header names
        header = category_headers[cat_idx]
        width = max(12, min(20, len(header) + 2))
        ws.column_dimensions[col_letter].width = width

    # Total column width
    total_col_letter = get_column_letter(total_col)
    ws.column_dimensions[total_col_letter].width = 12

    # Create QA sheet
    qa_ws = wb.create_sheet(title="QA Report")
    write_qa_sheet(qa_ws, qa_report, category_headers)

    # Save the file with original filename prefix
    job_id_short = job_id[:8] if job_id else "output"
    if original_filename:
        output_path = output_dir / f"{original_filename}_Contracts_Forms_{job_id_short}.xlsx"
    else:
        output_path = output_dir / f"Contracts_Forms_{job_id_short}.xlsx"
    wb.save(output_path)
    wb.close()

    return str(output_path)


def write_qa_sheet(ws, qa_report: QAReport, category_headers: List[str] = None):
    """
    Write QA report data to a worksheet.

    Args:
        ws: The worksheet to write to
        qa_report: QA report data
        category_headers: List of category headers (for reference)
    """
    header_font = Font(bold=True, size=16)
    normal_font = Font(size=16)
    accounting_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    current_row = 1

    # Parsing statistics
    ws.cell(row=current_row, column=1, value="Parsing Statistics").font = header_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="Total Rows Seen:").font = normal_font
    ws.cell(row=current_row, column=2, value=qa_report.parse_meta.total_rows_seen).font = normal_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="Rows Parsed:").font = normal_font
    ws.cell(row=current_row, column=2, value=qa_report.parse_meta.rows_parsed).font = normal_font
    current_row += 1
    ws.cell(row=current_row, column=1, value="Rows Skipped (Missing Fields):").font = normal_font
    ws.cell(row=current_row, column=2, value=qa_report.parse_meta.rows_skipped_missing_fields).font = normal_font
    current_row += 2

    # Counts per category
    ws.cell(row=current_row, column=1, value="Counts Per Category").font = header_font
    current_row += 1
    for bucket, count in sorted(qa_report.counts_per_bucket.items()):
        ws.cell(row=current_row, column=1, value=bucket).font = normal_font
        ws.cell(row=current_row, column=2, value=count).font = normal_font
        # Mark auto-created categories
        if bucket not in ["EXT PRIME", "EXTERIOR", "EXTERIOR UA", "INTERIOR",
                          "ROLL WALLS FINAL", "TOUCH UP", "Q4 REVERSAL", "UNMAPPED",
                          "UNDERCOAT", "BASE SHOE"]:
            ws.cell(row=current_row, column=3, value="(auto-created)").font = normal_font
        current_row += 1
    current_row += 1

    # Auto-created categories section
    auto_created = [ex for ex in qa_report.unmapped_examples if "[AUTO-CREATED]" in str(ex.get("task_text", ""))]
    if auto_created:
        ws.cell(row=current_row, column=1, value="Auto-Created Categories").font = header_font
        current_row += 1
        for item in auto_created:
            header_name = str(item.get("task_text", "")).replace("[AUTO-CREATED] ", "")
            ws.cell(row=current_row, column=1, value=header_name).font = normal_font
            ws.cell(row=current_row, column=2, value=f"{item.get('count', 0)} rows").font = normal_font
            current_row += 1
            # Show examples
            examples = item.get("examples", [])
            for ex in examples[:3]:
                ws.cell(row=current_row, column=2, value=f"  - {ex[:80]}...").font = normal_font
                current_row += 1
        current_row += 1

    # Unmapped tasks (excluding auto-created marker entries)
    regular_unmapped = [ex for ex in qa_report.unmapped_examples if "[AUTO-CREATED]" not in str(ex.get("task_text", ""))]
    if regular_unmapped:
        ws.cell(row=current_row, column=1, value="Unmapped Task Examples").font = header_font
        current_row += 1
        for example in regular_unmapped[:20]:
            ws.cell(row=current_row, column=1, value=example["task_text"]).font = normal_font
            ws.cell(row=current_row, column=2, value=example["count"]).font = normal_font
            current_row += 1
        current_row += 1

    # Suspicious totals
    if qa_report.suspicious_totals:
        ws.cell(row=current_row, column=1, value="Suspicious Totals").font = header_font
        current_row += 1
        for item in qa_report.suspicious_totals[:20]:
            ws.cell(row=current_row, column=1, value=f"Lot {item['lot_block']}").font = normal_font
            ws.cell(row=current_row, column=2, value=f"Plan {item['plan']}").font = normal_font
            cell = ws.cell(row=current_row, column=3, value=item['total'])
            cell.number_format = accounting_format
            cell.font = normal_font
            ws.cell(row=current_row, column=4, value=item['reason']).font = normal_font
            current_row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30


def write_summary_excel_legacy(
    summary_rows: List[SummaryRow],
    qa_report: QAReport,
    job_id: str,
    phase: str = None,
    project_name: str = None,
    house_string: str = None,
    original_filename: str = None
) -> str:
    """
    Legacy function for backward compatibility with SummaryRow objects.

    Converts SummaryRow objects to dicts and uses fixed category headers.
    """
    # Fixed category headers for legacy mode
    category_headers = [
        "EXT PRIME", "EXTERIOR", "EXTERIOR UA", "INTERIOR",
        "ROLL WALLS FINAL", "TOUCH UP", "Q4 REVERSAL"
    ]

    # Convert SummaryRow objects to dicts
    rows_as_dicts = []
    for row in summary_rows:
        row_dict = {
            "lot_block": row.lot_block,
            "plan": row.plan,
            "EXT PRIME": row.ext_prime,
            "EXTERIOR": row.exterior,
            "EXTERIOR UA": row.exterior_ua,
            "INTERIOR": row.interior,
            "ROLL WALLS FINAL": row.roll_walls_final,
            "TOUCH UP": row.touch_up,
            "Q4 REVERSAL": row.q4_reversal,
            "total": row.total,
        }
        rows_as_dicts.append(row_dict)

    return write_summary_excel(
        rows_as_dicts,
        qa_report,
        job_id,
        category_headers,
        phase=phase,
        project_name=project_name,
        house_string=house_string,
        original_filename=original_filename
    )
