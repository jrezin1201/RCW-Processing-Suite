"""Parser for Lennar scheduled tasks Excel exports."""
from typing import List, Dict, Tuple, Optional, Any
from datetime import datetime
import os
import logging
import re

from app.models.schemas import ParsedRow, QAMeta

logger = logging.getLogger(__name__)


def extract_project_name_from_b3(ws) -> Optional[str]:
    """
    Extract project name from cell B3 or B2 (between first and second hyphen).
    Checks B2 first, then B3 as fallback.

    Args:
        ws: Worksheet object (openpyxl) or data list (pandas)

    Returns:
        Project name string, or None if not found
    """
    # For openpyxl worksheet
    if hasattr(ws, 'cell'):
        # Try B2 first (some files have data one row up)
        for row in [2, 3]:
            cell_value = ws.cell(row=row, column=2).value  # B2 or B3
            if cell_value:
                # Split by hyphens and get the second part (between first and second hyphen)
                parts = str(cell_value).split('-')
                if len(parts) >= 2:
                    project_name = parts[1].strip()
                    logger.info(f"Found project name in B{row}: {project_name}")
                    return project_name
    return None


def extract_house_string_from_b5(ws) -> Optional[str]:
    """
    Extract house string from cell B5 or B4 (text after "PH## - ").
    Checks B4 first, then B5 as fallback.

    Args:
        ws: Worksheet object (openpyxl) or data list (pandas)

    Returns:
        House string, or None if not found
    """
    # For openpyxl worksheet
    if hasattr(ws, 'cell'):
        # Try B4 first, then B5 (some files have data one row up)
        for row in [4, 5]:
            cell_value = ws.cell(row=row, column=2).value  # B4 or B5
            if cell_value:
                # Find the pattern "PH## - " and extract everything after it
                import re
                match = re.search(r'PH\d{2}\s*-\s*(.+)', str(cell_value))
                if match:
                    house_string = match.group(1).strip()
                    logger.info(f"Found house string in B{row}: {house_string}")
                    return house_string
    return None


def extract_phase_from_b5(ws) -> Optional[str]:
    """
    Extract phase number from cell B5 or B4 (PHxx pattern).
    Checks B4 first, then B5 as fallback.

    Args:
        ws: Worksheet object (openpyxl) or data list (pandas)

    Returns:
        Phase number as string (without leading zeros), or None if not found
    """
    # For openpyxl worksheet
    if hasattr(ws, 'cell'):
        # Try B4 first, then B5 (some files have data one row up)
        for row in [4, 5]:
            cell_value = ws.cell(row=row, column=2).value  # B4 or B5
            if cell_value:
                phase_pattern = re.compile(r'PH(\d{2})', re.IGNORECASE)
                match = phase_pattern.search(str(cell_value))
                if match:
                    phase_num = match.group(1).lstrip('0')  # Remove leading zeros
                    logger.info(f"Found phase {phase_num} from {match.group(0)} in B{row}")
                    return phase_num if phase_num else None
    return None


def extract_phase_from_column_d(ws, column_map: Dict[str, int], max_rows: int = 100) -> Optional[str]:
    """
    Extract phase number from column D by looking for PH pattern (e.g., PH07).

    Args:
        ws: Worksheet object (openpyxl) or data list (pandas)
        column_map: Column mapping dictionary
        max_rows: Maximum rows to scan

    Returns:
        Phase number as string (e.g., "7" from "PH07"), or None if not found
    """
    phase_pattern = re.compile(r'PH(\d{2})', re.IGNORECASE)

    # For openpyxl worksheet
    if hasattr(ws, 'iter_rows'):
        for row in ws.iter_rows(min_row=1, max_row=max_rows, values_only=True):
            if row and len(row) > 3:  # Column D is index 3
                cell_value = str(row[3]) if row[3] else ""
                match = phase_pattern.search(cell_value)
                if match:
                    phase_num = match.group(1).lstrip('0')  # Remove leading zeros
                    logger.info(f"Found phase {phase_num} from {match.group(0)} in column D")
                    return phase_num if phase_num else None

    return None


def extract_phase_from_pandas_data(data: List, max_rows: int = 100) -> Optional[str]:
    """
    Extract phase number from column D in pandas data.

    Args:
        data: List of row data
        max_rows: Maximum rows to scan

    Returns:
        Phase number as string, or None if not found
    """
    phase_pattern = re.compile(r'PH(\d{2})', re.IGNORECASE)

    rows_to_check = min(max_rows, len(data))
    for row in data[:rows_to_check]:
        if row and len(row) > 3:  # Column D is index 3
            cell_value = str(row[3]) if row[3] else ""
            match = phase_pattern.search(cell_value)
            if match:
                phase_num = match.group(1).lstrip('0')  # Remove leading zeros
                logger.info(f"Found phase {phase_num} from {match.group(0)} in column D")
                return phase_num if phase_num else None

    return None


def parse_lennar_export(filepath: str) -> Tuple[List[ParsedRow], QAMeta, Optional[str], Optional[str], Optional[str]]:
    """
    Parse a Lennar scheduled tasks Excel export (supports both .xls and .xlsx).

    Args:
        filepath: Path to the Excel file

    Returns:
        Tuple of (parsed rows, QA metadata, phase number, project name, house string)
    """
    # Try to detect format by actually attempting to parse
    # This handles cases where the extension doesn't match the actual format
    try:
        # Try modern format first (xlsx)
        logger.info(f"Attempting to parse {filepath} as .xlsx format")
        return parse_with_openpyxl(filepath)
    except Exception as e:
        logger.info(f"Failed with openpyxl (likely .xls format): {str(e)[:100]}")
        # Fallback to pandas for old format
        try:
            logger.info(f"Attempting to parse {filepath} as .xls format with pandas")
            return parse_with_pandas(filepath)
        except Exception as e2:
            logger.error(f"Failed with both parsers: {e2}")
            # Try one more time with pandas using None engine (auto-detect)
            import pandas as pd
            try:
                df = pd.read_excel(filepath, engine=None)
                logger.info(f"Successfully read with pandas auto-detect, processing...")
                return parse_with_pandas_df(df)
            except Exception as e3:
                logger.error(f"All parsing attempts failed: {e3}")
                raise Exception(f"Could not parse file as either .xls or .xlsx format: {str(e)[:200]}")


def parse_with_openpyxl(filepath: str) -> Tuple[List[ParsedRow], QAMeta, Optional[str], Optional[str], Optional[str]]:
    """Parse using openpyxl for .xlsx files."""
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # Extract metadata
    project_name = extract_project_name_from_b3(ws)
    phase = extract_phase_from_b5(ws)  # Try B5 first
    if not phase:  # Fallback to column D if not found in B5
        phase = extract_phase_from_column_d(ws, {}, max_rows=100)
    house_string = extract_house_string_from_b5(ws)

    # Find header row
    header_row_idx, column_map = find_header_row_openpyxl(ws)

    if header_row_idx is None:
        return [], QAMeta(total_rows_seen=0, rows_parsed=0, rows_skipped_missing_fields=0), phase, project_name, house_string

    # Parse data rows
    parsed_rows = []
    qa_meta = QAMeta()
    consecutive_blank_rows = 0
    max_consecutive_blanks = 30

    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=True), start=header_row_idx + 1):
        qa_meta.total_rows_seen += 1

        # Check if row is effectively blank
        if is_row_blank(row):
            consecutive_blank_rows += 1
            if consecutive_blank_rows >= max_consecutive_blanks:
                break
            continue

        consecutive_blank_rows = 0

        # Parse the row
        parsed_row = parse_row(row, column_map)

        # Skip if missing required fields
        if not parsed_row.lot_block or not parsed_row.task_text:
            qa_meta.rows_skipped_missing_fields += 1
            continue

        parsed_rows.append(parsed_row)
        qa_meta.rows_parsed += 1

    wb.close()
    return parsed_rows, qa_meta, phase, project_name, house_string


def parse_with_pandas(filepath: str) -> Tuple[List[ParsedRow], QAMeta, Optional[str], Optional[str], Optional[str]]:
    """Parse using pandas for .xls files (old format)."""
    import pandas as pd

    # Read the Excel file with pandas (try xlrd first for .xls)
    try:
        df = pd.read_excel(filepath, engine='xlrd')
    except:
        # Fallback to auto-detect
        df = pd.read_excel(filepath, engine=None)

    return parse_with_pandas_df(df)


def parse_with_pandas_df(df) -> Tuple[List[ParsedRow], QAMeta, Optional[str], Optional[str], Optional[str]]:
    """Parse a pandas DataFrame."""
    import pandas as pd

    # Convert DataFrame to list of lists for compatibility
    data = df.values.tolist()
    headers = df.columns.tolist()

    # Extract metadata from specific cells
    project_name = None
    house_string = None
    phase = None

    # Try to extract from specific cells (B2/B3 and B4/B5)
    if len(data) >= 5:
        # Try B2 first, then B3 for project name (row 1, then row 2 in 0-indexed data)
        for row_idx, row_name in [(1, 'B2'), (2, 'B3')]:
            if not project_name and len(data) > row_idx and len(data[row_idx]) >= 2:
                cell_value = data[row_idx][1]  # Column B
                if cell_value and pd.notna(cell_value):
                    parts = str(cell_value).split('-')
                    if len(parts) >= 2:
                        project_name = parts[1].strip()
                        logger.info(f"Found project name in {row_name}: {project_name}")
                        break

        # Try B4 first, then B5 for phase and house string (row 3, then row 4 in 0-indexed data)
        for row_idx, row_name in [(3, 'B4'), (4, 'B5')]:
            if not phase and len(data) > row_idx and len(data[row_idx]) >= 2:
                cell_value = data[row_idx][1]  # Column B
                if cell_value and pd.notna(cell_value):
                    # Extract phase
                    phase_pattern = re.compile(r'PH(\d{2})', re.IGNORECASE)
                    match = phase_pattern.search(str(cell_value))
                    if match:
                        phase = match.group(1).lstrip('0')
                        logger.info(f"Found phase {phase} from {match.group(0)} in {row_name}")

                    # Extract house string
                    if not house_string:
                        house_match = re.search(r'PH\d{2}\s*-\s*(.+)', str(cell_value))
                        if house_match:
                            house_string = house_match.group(1).strip()
                            logger.info(f"Found house string in {row_name}: {house_string}")

                    if phase and house_string:
                        break

    # Fallback to column D for phase if not found
    if not phase:
        phase = extract_phase_from_pandas_data(data, max_rows=100)

    # Find header row in data
    header_row_idx, column_map = find_header_row_pandas(headers, data)

    if header_row_idx is None:
        return [], QAMeta(total_rows_seen=0, rows_parsed=0, rows_skipped_missing_fields=0), phase, project_name, house_string

    # Parse data rows
    parsed_rows = []
    qa_meta = QAMeta()
    consecutive_blank_rows = 0
    max_consecutive_blanks = 30

    # Start from the row after headers
    start_idx = 0 if header_row_idx == -1 else header_row_idx

    for row_idx, row in enumerate(data[start_idx:], start=start_idx):
        qa_meta.total_rows_seen += 1

        # Check if row is effectively blank
        if is_row_blank_pandas(row):
            consecutive_blank_rows += 1
            if consecutive_blank_rows >= max_consecutive_blanks:
                break
            continue

        consecutive_blank_rows = 0

        # Parse the row
        parsed_row = parse_row(row, column_map)

        # Skip if missing required fields
        if not parsed_row.lot_block or not parsed_row.task_text:
            qa_meta.rows_skipped_missing_fields += 1
            continue

        parsed_rows.append(parsed_row)
        qa_meta.rows_parsed += 1

    return parsed_rows, qa_meta, phase, project_name, house_string


def find_header_row_openpyxl(ws) -> Tuple[Optional[int], Dict[str, int]]:
    """Find header row for openpyxl worksheet."""
    from openpyxl.worksheet.worksheet import Worksheet

    required_headers = ["lot/block", "plan", "task", "task start date"]
    max_rows_to_check = 50

    for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows_to_check, values_only=True), start=1):
        if row is None:
            continue

        # Convert row to lowercase string list for comparison
        row_lower = [str(cell).lower().strip() if cell else "" for cell in row]

        # Check if all required headers are present
        found_headers = {}
        for header in required_headers:
            for col_idx, cell_value in enumerate(row_lower):
                if header in cell_value:
                    found_headers[header] = col_idx
                    break

        # If we found all required headers, build the column map
        if len(found_headers) == len(required_headers):
            column_map = build_column_map(row)
            return row_idx, column_map

    return None, {}


def find_header_row_pandas(headers: List, data: List) -> Tuple[Optional[int], Dict[str, int]]:
    """Find header row for pandas data."""
    required_headers = ["lot/block", "plan", "task", "task start date"]

    # First check if headers are in the column names
    headers_lower = [str(h).lower().strip() if h else "" for h in headers]
    found_headers = {}

    for header in required_headers:
        for col_idx, cell_value in enumerate(headers_lower):
            if header in cell_value:
                found_headers[header] = col_idx
                break

    if len(found_headers) == len(required_headers):
        column_map = build_column_map(headers)
        return -1, column_map  # -1 indicates headers are in column names

    # Otherwise search in data rows
    max_rows_to_check = min(50, len(data))

    for row_idx in range(max_rows_to_check):
        row = data[row_idx]
        if row is None:
            continue

        row_lower = [str(cell).lower().strip() if cell else "" for cell in row]
        found_headers = {}

        for header in required_headers:
            for col_idx, cell_value in enumerate(row_lower):
                if header in cell_value:
                    found_headers[header] = col_idx
                    break

        if len(found_headers) == len(required_headers):
            column_map = build_column_map(row)
            return row_idx, column_map

    return None, {}


def find_header_row(ws) -> Tuple[Optional[int], Dict[str, int]]:
    """Legacy function - redirects to openpyxl version."""
    return find_header_row_openpyxl(ws)


def is_row_blank_pandas(row: list) -> bool:
    """Check if a pandas row is effectively blank."""
    import pandas as pd
    if row is None:
        return True
    return all(pd.isna(cell) or str(cell).strip() == "" for cell in row)


def build_column_map(header_row: tuple) -> Dict[str, int]:
    """
    Build a mapping of column names to indices.

    Args:
        header_row: The header row values

    Returns:
        Dictionary mapping column names to indices
    """
    column_map = {}

    for idx, header in enumerate(header_row):
        if header is None:
            continue

        header_lower = str(header).lower().strip()

        # Map headers to standardized names
        if "lot" in header_lower and "block" in header_lower:
            column_map["lot_block"] = idx
        elif "plan" in header_lower:
            column_map["plan"] = idx
        elif "elevation" in header_lower:
            column_map["elevation"] = idx
        elif "swing" in header_lower:
            column_map["swing"] = idx
        elif "task start date" in header_lower:
            column_map["task_start_date"] = idx
        elif "task" in header_lower and "date" not in header_lower:
            column_map["task_text"] = idx
        elif "subtotal" in header_lower:
            column_map["subtotal"] = idx
        elif "tax" in header_lower:
            column_map["tax"] = idx
        elif "total" in header_lower and "subtotal" not in header_lower:
            column_map["total"] = idx

    return column_map


def is_row_blank(row: tuple) -> bool:
    """
    Check if a row is effectively blank.

    Args:
        row: Row values

    Returns:
        True if row is blank
    """
    if row is None:
        return True
    return all(cell is None or str(cell).strip() == "" for cell in row)


def parse_row(row: tuple, column_map: Dict[str, int]) -> ParsedRow:
    """
    Parse a single data row.

    Args:
        row: Row values
        column_map: Column name to index mapping

    Returns:
        ParsedRow object
    """
    def get_value(key: str) -> Any:
        """Get value from row using column map."""
        idx = column_map.get(key)
        if idx is not None and idx < len(row):
            return row[idx]
        return None

    def parse_money(value: Any) -> Optional[float]:
        """Parse money value to float."""
        if value is None:
            return None

        # If it's already a number
        if isinstance(value, (int, float)):
            return float(value)

        # Convert to string and clean
        value_str = str(value).strip()
        # Remove currency symbols and commas
        value_str = value_str.replace("$", "").replace(",", "").strip()

        try:
            return float(value_str)
        except (ValueError, TypeError):
            return None

    def parse_date(value: Any) -> Optional[datetime]:
        """Parse date value."""
        if value is None:
            return None

        # If it's already a datetime
        if isinstance(value, datetime):
            return value

        # Try to parse string date
        try:
            # Common date formats
            date_str = str(value).strip()
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"]:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
        except:
            pass

        return None

    return ParsedRow(
        lot_block=str(get_value("lot_block")).strip() if get_value("lot_block") else None,
        plan=str(get_value("plan")).strip() if get_value("plan") else None,
        elevation=str(get_value("elevation")).strip() if get_value("elevation") else None,
        swing=str(get_value("swing")).strip() if get_value("swing") else None,
        task_start_date=parse_date(get_value("task_start_date")),
        task_text=str(get_value("task_text")).strip() if get_value("task_text") else None,
        subtotal=parse_money(get_value("subtotal")),
        tax=parse_money(get_value("tax")),
        total=parse_money(get_value("total"))
    )