from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Tuple

import openpyxl


RATE_PER_HOUR = 0.75
JOB_RE = re.compile(r"^\s*(\d{4})\b")


@dataclass
class RowOut:
    job_number: str
    hours: float
    dollars: float


def _to_float(x) -> float:
    if x is None:
        return 0.0
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "" or s == "-":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def compute_job_costs_from_xlsx(file_bytes: bytes) -> List[RowOut]:
    """
    Reads the first sheet of an XLSX and computes:
      - JobNumber: first 4 digits at start of Location
      - Hours: Total hours from col L (or M as fallback)
      - Dollars: Hours * 0.75

    Supports two formats:
    1. Standard: Job in B, Employee in D, Hours in L
    2. RC Wendt: Job in A, Employee in C, Hours in L

    First tries to find "Location Total" rows, if none found, sums all rows with valid job numbers.
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # Try to detect format by looking for job numbers
    # Check if Column A has job numbers (RC Wendt format)
    job_in_a = False
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        if row and len(row) > 0:
            val = str(row[0] if row[0] else "").strip()
            if JOB_RE.match(val):
                job_in_a = True
                break

    # Set columns based on detected format
    if job_in_a:
        # RC Wendt format: A=job, C=employee, L=hours
        COL_LOCATION = 1      # A (1-based)
        COL_EMPLOYEE = 3      # C (1-based)
    else:
        # Standard format: B=job, D=employee, L=hours
        COL_LOCATION = 2      # B (1-based)
        COL_EMPLOYEE = 4      # D (1-based)

    COL_TOTAL_L = 12          # L (always)
    COL_TOTAL_M = 13          # M (fallback)

    totals: Dict[str, float] = {}

    def get_cell(row, col_1_based):
        """Safely get cell value from row."""
        idx = col_1_based - 1
        return row[idx] if row and len(row) > idx else None

    def get_hours(row) -> float:
        """Get hours from row, prefer L, fallback to M."""
        # Prefer L, fallback to M
        h = _to_float(get_cell(row, COL_TOTAL_L))
        if h > 0:
            return h
        h2 = _to_float(get_cell(row, COL_TOTAL_M))
        return h2

    # Pass 1: try to find subtotal rows ("Location Total") anywhere in the row
    found_total_rows = 0
    current_job = None  # Track current job number for Location Total rows

    for row in ws.iter_rows(values_only=True):
        # Check if this row has a job number
        loc = get_cell(row, COL_LOCATION)
        loc_s = str(loc).strip() if loc is not None else ""
        m = JOB_RE.match(loc_s)
        if m:
            current_job = m.group(1)  # Update current job number

        # Look for "location" and "total" in employee cell OR anywhere in row
        emp = get_cell(row, COL_EMPLOYEE)
        emp_s = str(emp).strip().lower() if emp is not None else ""

        # Check entire row for location + total pattern
        row_text = " ".join([str(x).strip().lower() for x in row if x is not None])
        is_total_row = ("location" in emp_s and "total" in emp_s) or \
                       ("location" in row_text and "total" in row_text)

        if not is_total_row:
            continue

        # Use current job number for total rows
        if not current_job:
            continue

        hours = get_hours(row)
        if hours <= 0:
            continue

        totals[current_job] = totals.get(current_job, 0.0) + hours
        found_total_rows += 1

    # Pass 2 (fallback): if no total rows found, sum all employee rows by job
    if found_total_rows == 0:
        totals = {}
        for row in ws.iter_rows(values_only=True):
            loc = get_cell(row, COL_LOCATION)
            loc_s = str(loc).strip() if loc is not None else ""
            m = JOB_RE.match(loc_s)
            if not m:
                continue

            hours = get_hours(row)
            if hours <= 0:
                continue

            job = m.group(1)
            totals[job] = totals.get(job, 0.0) + hours

    # Build output rows
    out: List[RowOut] = []
    for job, hours in sorted(totals.items(), key=lambda t: int(t[0])):
        dollars = round(hours * RATE_PER_HOUR, 2)
        out.append(RowOut(job_number=job, hours=round(hours, 2), dollars=dollars))

    return out


def build_output_workbook(rows: List[RowOut]) -> bytes:
    """
    Creates an XLSX output file as bytes with columns:
    JobNumber | Hours | Dollars

    Includes totals at the top.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GasAndRig"

    # Calculate totals
    total_hours = sum(r.hours for r in rows)
    total_dollars = sum(r.dollars for r in rows)

    # Add summary at top
    ws.append(["SUMMARY", "", ""])
    ws.append(["Total Hours:", round(total_hours, 2), ""])
    ws.append(["Total Dollars:", round(total_dollars, 2), ""])  # Store as number, not string
    ws.append(["", "", ""])  # Empty row for spacing

    # Add headers
    ws.append(["JobNumber", "Dollars", "Hours"])

    # Add data rows
    for r in rows:
        ws.append([int(r.job_number), r.dollars, r.hours])  # Convert job_number to int

    # Formatting
    # Bold the summary labels and headers
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws["A2"].font = openpyxl.styles.Font(bold=True)
    ws["A3"].font = openpyxl.styles.Font(bold=True)

    # Bold and style the summary values
    ws["B2"].font = openpyxl.styles.Font(bold=True, size=12)
    ws["B3"].font = openpyxl.styles.Font(bold=True, size=12)

    # Apply number format to Total Hours (2 decimal places)
    ws["B2"].number_format = '#,##0.00'

    # Apply currency format to Total Dollars
    ws["B3"].number_format = '"$"#,##0.00'

    # Bold headers row
    for cell in ws[5]:  # Headers are now in row 5
        if cell.value:
            cell.font = openpyxl.styles.Font(bold=True)

    # Apply formatting to all data rows
    for row_num in range(6, 6 + len(rows)):  # Data starts at row 6
        ws.cell(row=row_num, column=2).number_format = '"$"#,##0.00'  # Column B (Dollars)
        ws.cell(row=row_num, column=3).number_format = '#,##0.00'  # Column C (Hours)

    # Set column widths
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()