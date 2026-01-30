#!/usr/bin/env python3
"""Test that project name is correctly extracted from B3 and displayed in output."""
import openpyxl
from openpyxl import Workbook

# Create test data matching the format from the image
wb = Workbook()
ws = wb.active

# Set B3 with the exact format from the image
ws['B3'] = "OCH - Arches - 1157451"  # Should extract "Arches"
ws['B5'] = "PH07 - HS 44-49, 144-142"

# Add some test data
test_data = [
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["", "", "", ""],
    ["Lot/Block", "Plan", "Elevation", "Task", "Task Start Date", "Subtotal"],
    ["44", "2B", "B", "Prime Exterior", "2024-01-01", "1543.50"],
]

for row_idx, row_data in enumerate(test_data[6:], 7):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

test_filename = "test_project_extraction_input.xlsx"
wb.save(test_filename)
wb.close()

print("=" * 60)
print("PROJECT NAME EXTRACTION TEST")
print("=" * 60)
print(f"\nInput file created: {test_filename}")
print(f"Cell B3 contains: 'OCH - Arches - 1157451'")
print(f"Expected extraction: 'Arches'")

# Process the file
from app.services.parser_lennar import parse_lennar_export
from app.services.aggregator import aggregate_data
from app.services.excel_writer import write_summary_excel

print("\n1. EXTRACTION PHASE:")
parsed_rows, qa_meta, phase, project_name, house_string = parse_lennar_export(test_filename)

print(f"   - Extracted project name: '{project_name}'")
print(f"   - Extracted phase: '{phase}'")
print(f"   - Extracted house string: '{house_string}'")

if project_name == "Arches":
    print("   ✅ Project name extracted correctly!")
else:
    print(f"   ❌ Expected 'Arches', got '{project_name}'")

# Generate the output
print("\n2. OUTPUT GENERATION:")
summary_rows, qa_report = aggregate_data(parsed_rows, qa_meta)
output_path = write_summary_excel(
    summary_rows, qa_report, "test_project",
    phase=phase, project_name=project_name, house_string=house_string
)

print(f"   - Output file: {output_path}")

# Check the output file
wb = openpyxl.load_workbook(output_path)
ws = wb.active

print("\n3. OUTPUT VERIFICATION:")
b1_value = ws["B1"].value
g1_value = ws["G1"].value
h1_value = ws["H1"].value

print(f"   - Cell B1 (Project Name): '{b1_value}'")
print(f"   - Cell G1 (Phase): '{g1_value}'")
print(f"   - Cell H1 (House String): '{h1_value}'")

# Check merged cells
merged_ranges = ws.merged_cells.ranges
is_merged = any("B1:C1" in str(r) for r in merged_ranges)
print(f"   - B1:C1 merged: {'✅ Yes' if is_merged else '❌ No'}")

# Check if project name is displayed correctly
if "Arches" in str(b1_value):
    print("\n   ✅ Project name 'Arches' is displayed in the output!")
else:
    print(f"\n   ❌ Project name not found in output. B1 contains: '{b1_value}'")

# Check phase display
if "7" in str(g1_value):
    print("   ✅ Phase 7 is displayed correctly!")
else:
    print(f"   ❌ Phase not displayed correctly. G1 contains: '{g1_value}'")

wb.close()

print("\n" + "=" * 60)
print("TEST COMPLETE")
print("=" * 60)

# Clean up
import os
os.remove(test_filename)
if os.path.exists(output_path):
    os.remove(output_path)